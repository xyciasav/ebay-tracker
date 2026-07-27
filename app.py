import os
import hmac
import secrets
import base64
from datetime import datetime, timedelta
from pathlib import Path
from sqlalchemy import func, case, text, or_
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, Response, abort, session, jsonify
from werkzeug.utils import secure_filename
from PIL import Image, ImageOps
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_httpauth import HTTPBasicAuth
from functools import wraps
import csv
import json
import io, re
import html
import socket
import uuid
import urllib.request
import urllib.parse
import urllib.error
import openpyxl
import math
from difflib import SequenceMatcher



from models import db, Item, ItemImage, ScannerWatchKeyword, AppSetting

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def parse_float(value: str):
    if value is None:
        return None
    v = value.strip()
    if v == "":
        return None
    v = v.replace("$", "").replace(",", "")
    try:
        return float(v)
    except ValueError:
        return None


def parse_date(value: str):
    if not value:
        return None
    v = value.strip()
    if not v:
        return None
    try:
        return datetime.strptime(v, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_int(value: str):
    if value is None:
        return None
    v = str(value).strip().replace(",", "")
    if not v:
        return None
    try:
        return int(float(v))
    except ValueError:
        return None


def get_distinct_values(model, column):
    rows = db.session.query(column).distinct().filter(column.isnot(None)).order_by(column).all()
    values = []
    for r in rows:
        if not r or r[0] is None:
            continue
        s = str(r[0]).strip()
        if s:
            values.append(s)
    return values


def process_image(path: str, max_size: int = 1600):
    """
    Shrinks huge phone photos and fixes sideways rotation using EXIF.
    Overwrites the file at 'path' with an optimized version.
    """
    try:
        # Verify before processing so an extension cannot masquerade as an image.
        with Image.open(path) as probe:
            probe.verify()

        img = Image.open(path)
        img = ImageOps.exif_transpose(img)  # auto-rotate correctly
        img.thumbnail((max_size, max_size))  # keep aspect ratio

        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")

        img.save(path, optimize=True, quality=85)

    except Exception as e:
        print(f"Image processing failed for {path}: {e}")
        return False

    return True


def _normalize_url(value: str):
    value = (value or "").strip()
    if not value:
        return None
    parsed = urllib.parse.urlparse(value)
    if not parsed.scheme:
        value = f"https://{value}"
        parsed = urllib.parse.urlparse(value)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return None
    return urllib.parse.urlunparse(parsed)


def _ebay_item_url(item_number: str):
    item_number = (item_number or "").strip()
    return f"https://www.ebay.com/itm/{item_number}" if item_number else None


def _extract_ebay_item_number(url: str):
    url = _normalize_url(url)
    if not url or not _host_allowed_for_ebay_page(url):
        return None
    match = re.search(r"/itm/(?:[^/?#]+/)?(\d{8,})", urllib.parse.urlparse(url).path)
    return match.group(1) if match else None


def _sync_ebay_url_from_number(item: Item):
    item_number = (item.ebay_item_number or "").strip()
    if not item_number:
        return False
    canonical_url = _ebay_item_url(item_number)
    if item.ebay_item_url != canonical_url:
        item.ebay_item_url = canonical_url
        return True
    return False


def _host_allowed_for_ebay_page(url: str) -> bool:
    try:
        host = (urllib.parse.urlparse(url).hostname or "").lower()
    except Exception:
        return False
    return host == "ebay.com" or host.endswith(".ebay.com")


def _host_allowed_for_ebay_image(url: str) -> bool:
    try:
        host = (urllib.parse.urlparse(url).hostname or "").lower()
    except Exception:
        return False
    return host == "i.ebayimg.com" or host.endswith(".ebayimg.com")


def _host_is_private_or_local(url: str) -> bool:
    try:
        host = urllib.parse.urlparse(url).hostname
        if not host:
            return True
        infos = socket.getaddrinfo(host, None)
        import ipaddress
        for info in infos:
            ip = ipaddress.ip_address(info[4][0])
            if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved:
                return True
    except Exception:
        return True
    return False


def _fetch_url_bytes(url: str, max_bytes: int, timeout: int = 12):
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
            "Referer": "https://www.ebay.com/",
            "Upgrade-Insecure-Requests": "1",
        },
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        data = resp.read(max_bytes + 1)
        if len(data) > max_bytes:
            raise ValueError("Downloaded file is too large.")
        return data, resp.headers.get("Content-Type", "")


def _extract_og_image(html_text: str):
    patterns = [
        r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']',
        r'<meta[^>]+name=["\']twitter:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+name=["\']twitter:image["\']',
    ]
    for pattern in patterns:
        match = re.search(pattern, html_text, flags=re.IGNORECASE)
        if match:
            return html.unescape(match.group(1).strip())
    return None


def _save_image_from_url(item: Item, image_url: str, upload_folder: str):
    image_url = _normalize_url(image_url)
    if not image_url or _host_is_private_or_local(image_url):
        raise ValueError("The image URL is not allowed.")

    data, content_type = _fetch_url_bytes(image_url, max_bytes=10 * 1024 * 1024)
    ext = "jpg"
    lowered = content_type.lower()
    if "png" in lowered:
        ext = "png"
    elif "webp" in lowered:
        ext = "webp"
    elif "jpeg" in lowered or "jpg" in lowered:
        ext = "jpg"
    else:
        path_ext = Path(urllib.parse.urlparse(image_url).path).suffix.lower().lstrip(".")
        if path_ext in ALLOWED_EXTENSIONS:
            ext = "jpg" if path_ext == "jpeg" else path_ext

    stored_name = f"SKU{item.sku}_ebay_{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}_{uuid.uuid4().hex[:8]}.{ext}"
    save_path = os.path.join(upload_folder, stored_name)
    with open(save_path, "wb") as out:
        out.write(data)

    if not process_image(save_path):
        if os.path.exists(save_path):
            os.remove(save_path)
        raise ValueError("eBay image was downloaded, but it was not a valid image.")

    db.session.add(ItemImage(item_sku=item.sku, filename=stored_name))
    return stored_name


def _prepare_shelf_triage_image(file_storage, max_size: int = 1280):
    data = file_storage.read()
    if not data:
        raise ValueError("Choose a shelf photo first.")
    if len(data) > 12 * 1024 * 1024:
        raise ValueError("Shelf photo is too large. Try a closer crop or a smaller photo.")

    try:
        with Image.open(io.BytesIO(data)) as img:
            img = ImageOps.exif_transpose(img)
            img.thumbnail((max_size, max_size))
            if img.mode != "RGB":
                img = img.convert("RGB")

            out = io.BytesIO()
            img.save(out, format="JPEG", optimize=True, quality=82)
            encoded = base64.b64encode(out.getvalue()).decode("ascii")
            return f"data:image/jpeg;base64,{encoded}"
    except Exception:
        raise ValueError("That file does not look like a readable image.")


def _extract_openai_response_text(payload: dict):
    direct = payload.get("output_text")
    if isinstance(direct, str) and direct.strip():
        return direct.strip()

    pieces = []
    for output in payload.get("output", []) or []:
        for content in output.get("content", []) or []:
            text = content.get("text") or content.get("output_text")
            if isinstance(text, str) and text.strip():
                pieces.append(text.strip())
    return "\n".join(pieces).strip()


def _extract_chat_completion_text(payload: dict):
    choices = payload.get("choices", []) if isinstance(payload, dict) else []
    if not choices:
        return ""
    content = ((choices[0] or {}).get("message") or {}).get("content")
    if isinstance(content, str):
        return content.strip()
    if isinstance(content, list):
        pieces = []
        for part in content:
            if isinstance(part, dict) and isinstance(part.get("text"), str):
                pieces.append(part["text"].strip())
        return "\n".join(x for x in pieces if x).strip()
    return ""


def _vision_api_endpoint(base_url: str):
    base = (base_url or "").strip().rstrip("/")
    if not base:
        return None
    if base.endswith("/chat/completions") or base.endswith("/responses"):
        return base
    if base.endswith("/v1"):
        return f"{base}/chat/completions"
    return f"{base}/v1/chat/completions"


def _post_vision_json(endpoint: str, body: dict, headers: dict, timeout: int = 45):
    req = urllib.request.Request(
        endpoint,
        data=json.dumps(body).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _loads_json_object(text: str):
    text = (text or "").strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s*```$", "", text)

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start >= 0 and end > start:
            return json.loads(text[start:end + 1])
        raise


def _parse_shelf_triage_text_fallback(text: str):
    def clean_jsonish(value: str):
        value = (value or "").replace('\\"', '"')
        return re.sub(r"\s+", " ", value).strip()

    def extract_jsonish_bucket(key: str):
        key_match = re.search(rf'"{re.escape(key)}"\s*:\s*\[', text or "", flags=re.IGNORECASE)
        if not key_match:
            return []
        start = key_match.end()
        next_key = re.search(r'"\s*(?:focus_first|maybe_check|probably_skip|visible_text|summary)\s*"\s*:', (text or "")[start:], flags=re.IGNORECASE)
        segment = (text or "")[start:start + next_key.start()] if next_key else (text or "")[start:]
        items = []
        for obj in re.findall(r"\{[^{}]*\}", segment, flags=re.DOTALL):
            entry = {}
            for field in ("label", "why", "search_phrase", "confidence", "location", "evidence"):
                match = re.search(rf'"{field}"\s*:\s*"([^"]*)"', obj, flags=re.DOTALL)
                if match:
                    entry[field] = clean_jsonish(match.group(1))
            if entry.get("label"):
                entry.setdefault("search_phrase", entry["label"])
                entry.setdefault("why", "")
                entry.setdefault("confidence", "")
                items.append(entry)
        return items[:8]

    def clean_line(value: str):
        value = re.sub(r"^\s*[*\-•]+\s*", "", value or "")
        value = re.sub(r"^\s*\d+[.)]\s*", "", value)
        value = value.replace("**", "").replace("__", "").replace("`", "")
        return re.sub(r"\s+", " ", value).strip()

    def entry_from_line(line: str):
        line = clean_line(line)
        if not line or len(line) < 3:
            return None
        lowered = line.lower().replace("_", " ")
        if lowered.startswith(("focus first", "maybe check", "probably skip", "summary", "visible text")):
            return None

        label, reason = line, ""
        if ":" in line:
            label, reason = line.split(":", 1)
        elif " - " in line:
            label, reason = line.split(" - ", 1)

        label = clean_line(label).strip('"')
        reason = clean_line(reason)
        if re.fullmatch(r"item\s+\d+", label, flags=re.IGNORECASE) and reason:
            if "." in reason:
                label, reason = reason.split(".", 1)
            else:
                label, reason = reason, ""
            label = clean_line(label).strip('"')
            reason = clean_line(reason)
        label, reason = _split_triage_label_reason(label, reason)
        if not label:
            return None
        search_phrase = _clean_triage_search_phrase(label, label)
        return {
            "label": label[:140],
            "why": reason[:260],
            "search_phrase": search_phrase[:180],
            "confidence": "medium" if reason and re.search(r"\b(?:valuable|collectible|popular|series|set|board game|hardcover|edition)\b", reason, flags=re.IGNORECASE) else "",
        }

    buckets = {
        "focus_first": extract_jsonish_bucket("focus_first"),
        "maybe_check": extract_jsonish_bucket("maybe_check"),
        "probably_skip": extract_jsonish_bucket("probably_skip"),
    }
    current = None
    for raw_line in (text or "").splitlines():
        line = clean_line(raw_line)
        lowered = line.lower().replace("_", " ").strip(":")
        if not line:
            continue
        if "focus first" in lowered:
            current = "focus_first"
            continue
        if "maybe check" in lowered:
            current = "maybe_check"
            continue
        if "probably skip" in lowered:
            current = "probably_skip"
            continue
        if any(marker in lowered for marker in ("visible text", "final polish", "refining", "construct json", "extract visible")):
            current = None
            continue
        if current and len(buckets[current]) < 5 and re.match(r"^\s*(?:[*\-•]|\d+[.)])", raw_line):
            entry = entry_from_line(raw_line)
            if entry:
                buckets[current].append(entry)

    if not any(buckets.values()):
        return None

    for key, values in buckets.items():
        seen = set()
        deduped = []
        for item in values:
            label = (item.get("label") or "").lower()
            if not label or label in seen:
                continue
            seen.add(label)
            deduped.append(item)
        buckets[key] = deduped[:5]

    visible_text = []
    for quoted in re.findall(r'"([^"]{2,60})"', text or ""):
        if quoted.lower() not in {"summary", "focus_first", "maybe_check", "probably_skip", "visible_text"}:
            visible_text.append(quoted)

    return {
        "summary": "The vision model returned notes instead of strict JSON, so I pulled out the priority buckets from its response.",
        "focus_first": buckets["focus_first"],
        "maybe_check": buckets["maybe_check"],
        "probably_skip": buckets["probably_skip"],
        "visible_text": visible_text[:12],
    }


def _clean_triage_label(value: str):
    value = re.sub(r"\s+", " ", str(value or "")).strip()
    value = value.replace("“", '"').replace("”", '"').strip()
    value = value.replace('"', "")
    value = re.sub(r"\s*\((?:top|middle|bottom|left|right|second|third|fourth|shelf|bookshelf|check|inspect|too blurry|needs?)[^)]*\)\s*", " ", value, flags=re.IGNORECASE)
    value = re.sub(r"\s*/\s*hob$", " / Hobbit", value, flags=re.IGNORECASE)
    if value.count("(") > value.count(")"):
        value = value.replace("(", "").strip()
    value = re.sub(r"\s+", " ", value).strip(" .")
    value = value.strip('"')
    return value


def _split_triage_label_reason(label: str, reason: str):
    label = _clean_triage_label(label)
    reason = re.sub(r"\s+", " ", str(reason or "")).strip()
    reason = re.split(r"\b(?:wait|looking closer|or maybe|actually|ac[t]?\b)\b", reason, maxsplit=1, flags=re.IGNORECASE)[0].strip(" .")
    sentence_match = re.match(r'^(.{4,90}?)\.\s+(.+)$', label)
    if sentence_match and not reason:
        label = _clean_triage_label(sentence_match.group(1))
        reason = sentence_match.group(2).strip()
    return label, reason


def _clean_triage_search_phrase(value: str, fallback: str):
    value = _clean_triage_label(value or fallback)
    if ". " in value:
        value = value.split(". ", 1)[0]
    value = re.sub(r"\s*\((?:check|inspect|too blurry|needs?)[^)]*\)\s*", " ", value, flags=re.IGNORECASE)
    value = re.sub(r"\s*/\s*hob$", " / Hobbit", value, flags=re.IGNORECASE)
    value = re.sub(r"\s+", " ", value).strip(" .")
    return value or _clean_triage_label(fallback)


def _is_noise_shelf_triage_item(item: dict, bucket: str):
    label_value = str(item.get("label") or "").strip().lower()
    search_value = str(item.get("search_phrase") or "").strip().lower()
    text_value = " ".join([
        str(item.get("label") or ""),
        str(item.get("why") or ""),
        str(item.get("search_phrase") or ""),
    ]).lower()
    if not text_value.strip():
        return True

    non_item_labels = (
        "image analysis", "photo analysis", "analyze the image", "analysis of image",
        "photo description", "shelf photo", "shelf triage",
        "refine selections", "final answer", "analysis", "image description",
        "constructing the json", "drafting the json", "building the json",
        "json", "return json", "final json", "response json", "item", "visible item",
        "label", "search phrase", "confidence", "where", "evidence", "why", "location",
        "checking against constraints", "constraints", "no image analysis labels",
        "left", "right", "middle", "middle/right", "left/right", "far right", "far left",
        "top", "bottom", "left stack", "right stack", "middle stack", "top stack",
        "bottom stack", "left side", "right side", "middle section", "top section",
        "bottom section", "left section", "right section",
        "top right boxes", "bottom shelf red spines", "red spines", "spines",
        "left shelf", "right shelf", "middle shelf", "top shelf",
        "bottom shelf", "bookshelf", "bookcase", "tv stand", "entertainment center",
        "first shelf", "second shelf", "third shelf", "fourth shelf",
        "shelf 1", "shelf 2", "shelf 3", "shelf 4",
        "second shelf down", "third shelf down", "left shelf (bookshelf)",
    )
    if label_value in non_item_labels or search_value in non_item_labels:
        return True
    location_only_pattern = (
        r"(?:far\s+)?(?:left|right|middle|center|top|bottom)"
        r"(?:\s+(?:stack|side|section|area|pile|group|shelf|row|box|boxes))?"
    )
    if re.fullmatch(location_only_pattern, label_value) or re.fullmatch(location_only_pattern, search_value):
        return True
    if re.fullmatch(r"(?:left|right|middle|top|bottom|first|second|third|fourth)\s+shelf(?:\s+down)?(?:\s*\([^)]*\))?", label_value):
        return True
    if re.fullmatch(r"(?:left|right|middle|top|bottom|first|second|third|fourth)\s+shelf(?:\s+down)?", search_value):
        return True
    if re.search(r"\b(?:box|toy|item|object|thing)\s+on\s+(?:tv stand|shelf|table|stand)\b", label_value):
        return True
    if re.search(r"\b(?:box|toy|item|object|thing)\s+on\s+(?:tv stand|shelf|table|stand)\b", search_value):
        return True

    hard_skip_terms = (
        "paper", "papers", "clutter", "junk mail", "envelope", "stationery",
        "ruler", "rulers", "marker", "pen", "glue", "loose cable", "cables",
        "cords", "tangled wires", "wires", "accessories", "trash bin", "trash can", "garbage can",
        "generic books", "generic book", "random books", "loose books",
        "books/dvds", "books / dvds", "middle shelves", "shelf contents",
        "top right boxes", "bottom shelf red spines", "red spines",
    )
    if any(term in text_value for term in hard_skip_terms):
        return True
    if label_value in {"book", "books", "loose books"} or search_value in {"book", "books", "loose books"}:
        return True

    generic_terms = (
        "generic toy", "toy on", "yellow toy", "grey toy", "gray toy",
        "small toy", "toys", "misc toy", "unbranded toy",
        "generic box", "unknown box", "small box on", "small boxes on", "box on tv stand",
        "top right boxes",
        "disney box", "nickelodeon box",
    )
    value_clues = (
        "brand", "branded", "model", "upc", "sealed", "new in box", "nib",
        "vintage", "rare", "figure", "figurine", "action figure", "statue",
        "doll", "plush", "glass", "porcelain", "ceramic", "crystal", "marked",
        "marking", "signature", "maker", "stamp", "label", "tag", "bottom",
        "resin", "brass", "metal", "wooden", "handmade", "hand painted",
        "collectible", "character", "licensed", "anime", "manga", "comic",
        "star wars", "harry potter", "lord of the rings", "lotr", "hobbit",
        "zelda", "legend of zelda", "dungeons & dragons", "dungeons and dragons",
        "d&d", "dnd", "ad&d", "tsr", "pathfinder", "warhammer", "magic the gathering",
        "mtg", "pokemon", "yugioh", "dragon ball", "marvel", "dc comics", "bluey",
        "disney", "nintendo", "sony", "vizio", "lego", "funko", "barbie", "fisher-price",
    )
    if any(term in text_value for term in generic_terms) and not any(clue in text_value for clue in value_clues):
        return True
    if any(term in text_value for term in generic_terms) and any(negative in text_value for negative in ("no brand", "no character", "not visible", "can't identify", "cannot identify")):
        return True

    if bucket == "probably_skip" and any(term in text_value for term in ("generic", "low value", "clutter", "paper")):
        return True

    return False


def _normalize_shelf_triage(parsed):
    known_readable_titles = (
        "exploding kittens",
        "kids against maturity",
        "a christmas story the party game",
        "christmas story",
        "goosebumps",
        "legend of zelda",
        "lord of the rings",
        "the hobbit",
    )

    def is_vague_lookup(item):
        text_value = " ".join([
            str(item.get("label") or ""),
            str(item.get("search_phrase") or ""),
        ]).lower()
        vague_terms = (
            "toy", "toys", "box", "small box", "glass jar", "green plush",
            "plush toy", "white ceramic", "ceramic toilet", "star wars box",
            "anime figure", "blonde anime figure", "figure blonde",
            "book lot", "unknown", "inspect", "check tag",
            "maker mark", "bottom left", "tv stand",
        )
        return any(term in text_value for term in vague_terms)

    def visible_title_match(item, visible_blob: str):
        item_blob = _norm_title(" ".join([
            str(item.get("label") or ""),
            str(item.get("search_phrase") or ""),
        ]))
        for title in known_readable_titles:
            title_norm = _norm_title(title)
            if title_norm in visible_blob and title_norm in item_blob:
                return title
            if title_norm == "christmas story" and title_norm in visible_blob and "christmas story" in item_blob:
                return "A Christmas Story The Party Game"
        return ""

    def promote_visible_readable_titles(items, visible_blob: str):
        for item in items:
            matched_title = visible_title_match(item, visible_blob)
            if not matched_title:
                continue
            item["confidence"] = "high"
            if not item.get("evidence"):
                item["evidence"] = f"Detected text includes: {matched_title}."
            if not item.get("why") or "readable" not in item["why"].lower():
                item["why"] = (item.get("why") or "Readable title visible.")[:220]
                if "readable" not in item["why"].lower():
                    item["why"] = f"Readable title visible. {item['why']}".strip()[:260]
        return items

    def normalize_list(key):
        value = parsed.get(key) if isinstance(parsed, dict) else []
        if isinstance(value, str):
            value = [{"label": value}]
        if not isinstance(value, list):
            value = []

        items = []
        for entry in value:
            if isinstance(entry, str):
                entry = {"label": entry}
            if not isinstance(entry, dict):
                continue
            label = str(entry.get("label") or entry.get("item") or "Visible item").strip()
            reason = str(entry.get("why") or entry.get("reason") or "").strip()
            label, reason = _split_triage_label_reason(label, reason)
            if not label:
                continue
            search_phrase = _clean_triage_search_phrase(entry.get("search_phrase") or entry.get("query") or label, label)
            confidence = str(entry.get("confidence") or "").strip()
            if not confidence and re.search(r"\b(?:readable|specific|exact|popular|collectible|series|set|board game|hardcover|edition|valuable|value)\b", reason, flags=re.IGNORECASE):
                confidence = "medium"
            item = {
                "label": label[:140],
                "why": reason[:260],
                "search_phrase": search_phrase[:180],
                "confidence": confidence[:40],
                "location": str(entry.get("location") or entry.get("where") or "").strip()[:120],
                "evidence": str(entry.get("evidence") or entry.get("visual_evidence") or entry.get("seen") or "").strip()[:180],
            }
            if is_vague_lookup(item) and "high" in item["confidence"].lower():
                item["confidence"] = "medium"
            if _is_noise_shelf_triage_item(item, key):
                continue
            items.append(item)
            if len(items) >= 8:
                break
        return items

    def is_prompt_leak_text(value: str):
        normalized = _norm_title(value)
        if not normalized:
            return True
        leak_exact = {
            "watchlist", "sold history", "search phrase", "confidence", "label",
            "where", "evidence", "focus first", "inspect closer", "probably skip",
        }
        if normalized in leak_exact:
            return True
        if normalized.startswith(("watchlist ", "sold history ", "search phrase ")):
            return True
        return False

    visible_text = parsed.get("visible_text", []) if isinstance(parsed, dict) else []
    if isinstance(visible_text, str):
        visible_text = [visible_text]
    if not isinstance(visible_text, list):
        visible_text = []
    cleaned_visible_text = []
    seen_visible_text = set()
    for value in visible_text:
        cleaned = str(value).strip()[:80]
        normalized = _norm_title(cleaned)
        if not cleaned or is_prompt_leak_text(cleaned) or normalized in seen_visible_text:
            continue
        seen_visible_text.add(normalized)
        cleaned_visible_text.append(cleaned)
        if len(cleaned_visible_text) >= 12:
            break

    focus_first = normalize_list("focus_first")
    maybe_check = normalize_list("maybe_check")
    probably_skip = normalize_list("probably_skip")

    def add_visible_keyword_lead(label, search_phrase, why):
        existing = " ".join(
            item["label"].lower()
            for item in (focus_first + maybe_check + probably_skip)
            if item.get("label")
        )
        if label.lower() in existing or search_phrase.lower() in existing:
            return
        broad_keys = ("zelda", "dungeons", "d&d", "dnd", "tsr", "pathfinder", "warhammer", "magic", "mtg", "lord of the rings", "hobbit")
        if any(key in label.lower() and key in existing for key in broad_keys):
            return
        if len(focus_first) >= 5:
            maybe_check.append({
                "label": label,
                "why": why,
                "search_phrase": search_phrase,
                "confidence": "medium",
            })
        else:
            focus_first.append({
                "label": label,
                "why": why,
                "search_phrase": search_phrase,
                "confidence": "medium",
            })

    visible_blob = " ".join(cleaned_visible_text).lower()
    normalized_visible_blob = _norm_title(" ".join(cleaned_visible_text))
    focus_first = promote_visible_readable_titles(focus_first, normalized_visible_blob)
    maybe_check = promote_visible_readable_titles(maybe_check, normalized_visible_blob)
    if "one ring" not in visible_blob:
        for item in focus_first + maybe_check:
            item_text = f"{item.get('label', '')} {item.get('search_phrase', '')}".lower()
            if "one ring" in item_text:
                item["confidence"] = "low"
                if item.get("why"):
                    item["why"] = f"{item['why']} Verify the title first; OCR did not confirm it."
                else:
                    item["why"] = "Verify the title first; OCR did not confirm it."
    keyword_promotions = [
        ("Legend of Zelda", "Legend of Zelda", "Visible Zelda/Nintendo keyword; always worth checking."),
        ("Dungeons & Dragons books", "Dungeons & Dragons books", "Visible D&D/tabletop RPG books; check edition and publisher."),
        ("AD&D / TSR books", "AD&D TSR books", "Visible AD&D/TSR keyword; older RPG books can be valuable."),
        ("Pathfinder RPG books", "Pathfinder RPG books", "Visible tabletop RPG keyword; check edition/sourcebook."),
        ("Warhammer books/games", "Warhammer books", "Visible Warhammer keyword; check edition and condition."),
        ("Magic: The Gathering cards", "Magic The Gathering cards", "Visible MTG/card keyword; check sets and condition."),
        ("Lord of the Rings / Hobbit books", "Lord of the Rings Hobbit books", "Visible Tolkien keyword; boxed sets and older editions can be valuable."),
    ]
    for label, search_phrase, why in keyword_promotions:
        label_key = label.lower()
        if (
            ("zelda" in visible_blob and "zelda" in label_key)
            or (("dungeons" in visible_blob or "d&d" in visible_blob or "dnd" in visible_blob) and "dungeons" in label_key)
            or (("ad&d" in visible_blob or "tsr" in visible_blob) and "tsr" in label_key)
            or ("pathfinder" in visible_blob and "pathfinder" in label_key)
            or ("warhammer" in visible_blob and "warhammer" in label_key)
            or (("magic" in visible_blob or "mtg" in visible_blob) and "magic" in label_key)
            or (("lord of the rings" in visible_blob or "hobbit" in visible_blob or "lotr" in visible_blob) and "lord of the rings" in label_key)
        ):
            add_visible_keyword_lead(label, search_phrase, why)

    return {
        "summary": str(parsed.get("summary") or "Shelf triage complete.").strip()[:300] if isinstance(parsed, dict) else "Shelf triage complete.",
        "focus_first": focus_first[:8],
        "maybe_check": maybe_check[:8],
        "probably_skip": probably_skip[:5],
        "visible_text": cleaned_visible_text,
    }


_HISTORY_KEYWORD_STOPWORDS = {
    "the", "and", "for", "with", "from", "into", "that", "this", "your", "you",
    "new", "used", "lot", "bundle", "set", "book", "books", "game", "games",
    "card", "cards", "dvd", "bluray", "blu", "ray", "disc", "movie", "movies",
    "box", "boxed", "sealed", "complete", "vintage", "rare", "edition", "ed",
    "series", "collection", "official", "authentic", "toy", "toys", "figure",
    "figures", "shirt", "tshirt", "t", "men", "women", "kids", "size",
    "sun", "sand", "straw", "hat", "corded", "rope", "blue", "base", "beach",
    "black", "white", "red", "green", "yellow", "brown", "gray", "grey", "beige",
    "large", "small", "medium", "long", "short", "left", "right", "top", "bottom",
}

_HIGH_SIGNAL_SINGLE_HISTORY_TERMS = {
    "zelda", "nintendo", "pokemon", "yugioh", "warhammer", "pathfinder", "tsr",
    "dnd", "mtg", "lego", "funko", "barbie", "goosebumps", "tolkien", "hobbit",
    "lotr", "starwars", "star", "manga", "anime",
}


def _history_tokens(value: str):
    normalized = re.sub(r"[^a-z0-9&]+", " ", (value or "").lower())
    return [token for token in normalized.split() if len(token) >= 3 and token not in _HISTORY_KEYWORD_STOPWORDS]


def _scanner_history_keyword_promotions(limit: int = 250):
    try:
        rows = (
            Item.query
            .filter(Item.sold.is_(True), Item.sold_confirmed.is_(True), Item.canceled.is_(False))
            .order_by(Item.date_sold.desc().nullslast(), Item.updated_at.desc())
            .limit(limit)
            .all()
        )
    except Exception:
        return []

    promotions = {}
    for item in rows:
        title = _clean_triage_label(item.item_name or "")
        if not title:
            continue
        title_tokens = _history_tokens(title)
        candidates = set()

        if len(title_tokens) >= 2:
            candidates.add(" ".join(title_tokens[:5]))
        for token in title_tokens:
            if token in _HIGH_SIGNAL_SINGLE_HISTORY_TERMS:
                candidates.add(token)
        for size in (2, 3):
            for i in range(0, max(0, len(title_tokens) - size + 1)):
                candidates.add(" ".join(title_tokens[i:i + size]))

        # Do not promote broad eBay categories like "Computers Tablets Networking
        # Devices" or "Vinyl Records". Those produced noisy scanner leads because
        # a category can match a shelf without identifying a pick-up-worthy item.

        for candidate in candidates:
            candidate = candidate.strip()
            if " " not in candidate and candidate not in _HIGH_SIGNAL_SINGLE_HISTORY_TERMS:
                continue
            if len(candidate) < 4 and candidate not in {"dnd", "d&d", "tsr", "mtg"}:
                continue
            if candidate in _HISTORY_KEYWORD_STOPWORDS:
                continue
            record = promotions.setdefault(candidate, {
                "keyword": candidate,
                "label": title[:80],
                "search_phrase": title[:120],
                "count": 0,
                "profit": 0.0,
            })
            record["count"] += 1
            record["profit"] += float(item.profit or 0.0)

    sorted_promotions = sorted(
        promotions.values(),
        key=lambda r: (r["count"], r["profit"], len(r["keyword"])),
        reverse=True,
    )
    return sorted_promotions[:80]


def _apply_history_triage_boosts(triage: dict, promotions: list):
    if not promotions:
        return triage

    focus_first = list(triage.get("focus_first") or [])
    maybe_check = list(triage.get("maybe_check") or [])
    probably_skip = list(triage.get("probably_skip") or [])
    visible_text = list(triage.get("visible_text") or [])
    scan_blob = " ".join(
        visible_text +
        [str(item.get("label", "")) for item in focus_first + maybe_check + probably_skip] +
        [str(item.get("why", "")) for item in focus_first + maybe_check + probably_skip]
    ).lower()
    existing_blob = " ".join(str(item.get("label", "")) for item in focus_first + maybe_check + probably_skip).lower()

    for promo in promotions:
        keyword = promo["keyword"].lower()
        if len(keyword) < 4 and keyword not in {"dnd", "d&d", "tsr", "mtg"}:
            continue
        if keyword not in scan_blob:
            continue
        label = promo["label"]
        if label.lower() in existing_blob or keyword in existing_blob:
            continue
        lead = {
            "label": f"Sold-history match: {keyword.title()}",
            "why": f"You have sold related items before, for example: {label}.",
            "search_phrase": keyword if len(keyword.split()) >= 2 else promo["search_phrase"],
            "confidence": "medium",
        }
        if len(focus_first) < 5:
            focus_first.append(lead)
        elif len(maybe_check) < 5:
            maybe_check.append(lead)
        else:
            break
        existing_blob += " " + label.lower() + " " + keyword

    triage["focus_first"] = focus_first[:5]
    triage["maybe_check"] = maybe_check[:5]
    triage["probably_skip"] = probably_skip[:5]
    return triage


def _triage_item_priority(item: dict):
    label = str(item.get("label") or "").lower()
    why = str(item.get("why") or "").lower()
    search_phrase = str(item.get("search_phrase") or "").lower()
    confidence = str(item.get("confidence") or "").lower()
    text_blob = f"{label} {why} {search_phrase}"

    score = 0
    if label.startswith("watchlist match:"):
        score += 140
    if label.startswith("sold-history match:"):
        score += 95
    if "high" in confidence:
        score += 110
    elif "medium" in confidence:
        score += 60
    elif "low" in confidence:
        score += 20
    else:
        score += 35

    if re.search(r"\b(?:readable|specific|exact|model|upc|sealed|new in box|nib|boxed set|box set)\b", text_blob):
        score += 30
    if re.search(r"\b(?:zelda|nintendo|dungeons|dragons|d&d|dnd|ad&d|tsr|pathfinder|warhammer|magic the gathering|mtg|pokemon|yugioh|goosebumps|lord of the rings|hobbit|star wars|american girl|lego|funko)\b", text_blob):
        score += 22
    if re.search(r"\b(?:figure|figurine|statue|doll|plush|ceramic|porcelain|glass|crystal|maker mark|tag|label)\b", text_blob):
        score += 12

    if re.search(r"\b(?:generic|unknown|check tag|inspect bottom|maybe|could be|looks like|hard to read|not visible)\b", text_blob):
        score -= 18
    if re.search(r"\b(?:anime figure|figure|figurine|statue|doll|plush)\b", text_blob) and not re.search(r"\b(?:nendoroid|figma|funko|fisher-price|barbie|american girl|hasbro|mattel|bandai|banpresto|good smile|kotobukiya|hot toys|sideshow|pokemon|dragon ball|naruto|one piece|sailor moon|marvel|dc|star wars|disney|character|series|tag|label|maker|marked|number)\b", text_blob):
        score -= 30
    if re.search(r"\b(?:graphics card|gpu|video card|printer|scanner|console)\b", text_blob) and not re.search(r"\b(?:rtx|gtx|radeon|rx\s?\d|nvidia|amd|geforce|model|serial|upc|wii|playstation|xbox|switch|epson|canon|brother|hp)\b", text_blob):
        score -= 35
    if label in {"nintendo", "watchlist match: nintendo"} and re.search(r"\b(?:zelda|legend of zelda)\b", text_blob):
        score -= 60
    if label in {"book", "books", "tv", "white console", "green plush", "stuffed animals"}:
        score -= 40
    if re.fullmatch(r"(?:the\s+)?(?:large\s+)?center speaker", label) or label in {"headphones", "speaker", "speakers"}:
        score -= 45
    if label in {"why", "location", "label", "search phrase", "confidence", "where", "evidence"}:
        score -= 200

    return score


def _triage_dedupe_key(item: dict):
    value = _norm_title(" ".join([
        str(item.get("search_phrase") or item.get("label") or ""),
        str(item.get("evidence") or ""),
        str(item.get("why") or ""),
    ]))
    value = re.sub(r"^(watchlist match|sold history match|sold-history match)\s+", "", value)
    value = re.sub(r"\b(?:box|boxed|board|card|game|games|party|popular|used|visible|readable|high|confidence)\b", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    aliases = {
        "kids against maturity": "kids against maturity",
        "exploding kittens": "exploding kittens",
        "graphics gpu": "graphics card",
        "graphics card gpu": "graphics card",
        "pc video": "graphics card",
        "pc video card": "graphics card",
        "video card": "graphics card",
        "fender": "electric guitar" if re.search(r"\b(?:strat|stratocaster|guitar|telecaster)\b", value) else "fender",
        "squier": "electric guitar" if re.search(r"\b(?:strat|stratocaster|guitar|telecaster)\b", value) else "squier",
        "gibson": "electric guitar" if "guitar" in value else "gibson",
        "black stratocaster style guitar": "electric guitar",
        "black stratocaster guitar": "electric guitar",
        "black electric guitar": "electric guitar",
        "stratocaster guitar": "electric guitar",
        "electric guitar": "electric guitar",
        "the large center speaker": "center speaker",
        "large center speaker": "center speaker",
        "center speaker": "center speaker",
        "nintendo": "zelda" if "zelda" in value else "nintendo",
        "legend zelda": "zelda",
        "zelda": "zelda",
        "lord rings hobbit": "lord of the rings hobbit",
        "lord rings hob": "lord of the rings hobbit",
        "lord of rings hob": "lord of the rings hobbit",
        "lord of rings hobbit": "lord of the rings hobbit",
    }
    for needle, replacement in aliases.items():
        if needle in value:
            return replacement
    return value or _norm_title(item.get("label") or "")


def _dedupe_triage_items(items: list):
    kept = []
    seen = set()
    for item in _sort_triage_items(items):
        key = _triage_dedupe_key(item)
        if key in seen:
            continue
        seen.add(key)
        kept.append(item)
    return kept


def _sort_triage_items(items: list):
    return sorted(
        list(items or []),
        key=lambda item: (
            _triage_item_priority(item),
            str(item.get("label") or "").lower(),
        ),
        reverse=True,
    )


def _prioritize_triage_buckets(triage: dict):
    focus_first = list(triage.get("focus_first") or [])
    maybe_check = list(triage.get("maybe_check") or [])
    probably_skip = _sort_triage_items(triage.get("probably_skip") or [])

    combined = _dedupe_triage_items(focus_first + maybe_check)
    triage["focus_first"] = combined[:5]
    triage["maybe_check"] = combined[5:10]
    triage["probably_skip"] = probably_skip[:5]
    return triage


def _header_key(value: str):
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def _row_get(row: dict, *names):
    wanted = {_header_key(name) for name in names}
    for key, value in row.items():
        if _header_key(key) in wanted:
            return "" if value is None else str(value).strip()
    return ""


def _parse_scanner_watchlist_upload(file_storage):
    if not file_storage:
        raise ValueError("Choose a CSV or XLSX watchlist file first.")
    filename = (file_storage.filename or "").lower()
    raw = file_storage.read()
    if not raw:
        raise ValueError("Choose a CSV or XLSX watchlist file first.")

    rows = []
    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        values = ws.iter_rows(values_only=True)
        headers = [str(h or "").strip() for h in next(values, [])]
        for values_row in values:
            row = {headers[i]: values_row[i] if i < len(values_row) else "" for i in range(len(headers))}
            if any(str(v or "").strip() for v in row.values()):
                rows.append(row)
    elif filename.endswith(".csv"):
        text_stream = io.StringIO(raw.decode("utf-8-sig", errors="replace"))
        rows = list(csv.DictReader(text_stream))
    else:
        raise ValueError("Watchlist import supports .xlsx or .csv files.")

    entries = []
    for row in rows:
        category = _row_get(row, "category", "type", "department")
        brand = _row_get(row, "brand / company", "brand", "company", "keyword", "franchise", "item")
        item_keywords = _row_get(row, "item keywords", "keywords", "items", "search", "search phrase", "search_phrase")
        notes = _row_get(row, "notes", "note", "why", "details")
        priority = (_row_get(row, "priority") or "high").lower()
        comp_raw = (_row_get(row, "comp_link_allowed", "comp links", "show comps") or "yes").lower()
        comp_link_allowed = comp_raw not in {"no", "false", "0", "n"}

        keyword = _clean_triage_label(brand or item_keywords)
        if not keyword:
            continue
        label = _clean_triage_label(_row_get(row, "label", "display label") or keyword)
        search_phrase = _clean_triage_search_phrase(
            _row_get(row, "search_phrase", "search phrase", "search") or keyword,
            keyword,
        )
        if item_keywords and item_keywords != search_phrase:
            keyword_note = f"Keyword hints: {item_keywords}"
            notes = f"{notes}\n{keyword_note}".strip()
        if category:
            notes = f"Watchlist category: {category}. {notes}".strip()

        entries.append({
            "keyword": keyword[:160],
            "label": label[:180],
            "search_phrase": search_phrase[:220],
            "category": category[:140] if category else None,
            "priority": priority[:40] or "high",
            "comp_link_allowed": comp_link_allowed,
            "notes": notes,
        })
    return entries


def _scanner_watchlist_promotions(limit: int = 2500):
    try:
        rows = (
            ScannerWatchKeyword.query
            .filter(ScannerWatchKeyword.active.is_(True))
            .order_by(ScannerWatchKeyword.priority.asc(), ScannerWatchKeyword.keyword.asc())
            .limit(limit)
            .all()
        )
    except Exception:
        return []

    return [{
        "keyword": row.keyword,
        "label": row.label or row.keyword,
        "search_phrase": row.search_phrase or row.keyword,
        "category": row.category or "",
        "priority": row.priority or "high",
        "comp_link_allowed": bool(row.comp_link_allowed),
        "notes": row.notes or "",
    } for row in rows if row.keyword]


def _text_has_watch_keyword(text_blob: str, keyword: str):
    keyword_norm = _norm_title(keyword)
    text_norm = _norm_title(text_blob)
    if not keyword_norm or len(keyword_norm) < 3:
        return False
    return re.search(rf"\b{re.escape(keyword_norm)}\b", text_norm) is not None


def _visible_text_match(visible_text: list, keyword: str):
    for value in visible_text or []:
        if _text_has_watch_keyword(str(value), keyword):
            return str(value).strip()
    return ""


def _specific_watchlist_match_exists(scan_blob: str, keyword: str, promotions: list):
    keyword_norm = _norm_title(keyword)
    if keyword_norm == "nintendo":
        child_terms = {"zelda", "legend of zelda", "pokemon", "switch", "wii", "gamecube", "nes", "snes", "n64"}
        return any(
            _norm_title(promo.get("keyword") or "") in child_terms
            and _text_has_watch_keyword(scan_blob, promo.get("keyword") or "")
            for promo in promotions
        )
    return False


def _watchlist_match_has_context(keyword: str, matched_text: str, item_context: str):
    keyword_norm = _norm_title(keyword)
    matched_norm = _norm_title(matched_text)
    context_norm = _norm_title(item_context)
    combined = f"{matched_norm} {context_norm}"

    if keyword_norm == "jackson" and "michael jackson" in matched_norm:
        return False

    music_brands = {"fender", "gibson", "jackson", "squier", "ibanez", "epiphone"}
    if keyword_norm in music_brands:
        return bool(re.search(
            r"\b(?:guitar|strat|stratocaster|telecaster|bass|amp|amplifier|pedal|pickup|headstock|instrument)\b",
            combined,
        ))

    decor_brands = {"rae dunn", "pyrex", "fiestaware", "fiesta", "lenox", "lladro"}
    if keyword_norm in decor_brands:
        return bool(re.search(
            r"\b(?:mug|vase|bowl|plate|dish|canister|planter|pottery|ceramic|figurine|decor|decorative|marked|stamp|label)\b",
            combined,
        ))

    return True


def _apply_watchlist_triage_boosts(triage: dict, promotions: list):
    if not promotions:
        return triage

    focus_first = list(triage.get("focus_first") or [])
    maybe_check = list(triage.get("maybe_check") or [])
    probably_skip = list(triage.get("probably_skip") or [])
    visible_text = list(triage.get("visible_text") or [])
    scan_blob = " ".join(
        visible_text +
        [str(item.get("label", "")) for item in focus_first + maybe_check + probably_skip] +
        [str(item.get("why", "")) for item in focus_first + maybe_check + probably_skip]
    )
    item_context = " ".join(
        [str(item.get("label", "")) for item in focus_first + maybe_check + probably_skip] +
        [str(item.get("why", "")) for item in focus_first + maybe_check + probably_skip] +
        [str(item.get("evidence", "")) for item in focus_first + maybe_check + probably_skip] +
        [str(item.get("search_phrase", "")) for item in focus_first + maybe_check + probably_skip]
    )
    existing_blob = " ".join(str(item.get("label", "")) for item in focus_first + maybe_check + probably_skip).lower()

    priority_rank = {"high": 0, "medium": 1, "inspect": 2, "low": 3}
    for promo in sorted(promotions, key=lambda p: priority_rank.get((p.get("priority") or "high").lower(), 1)):
        keyword = promo.get("keyword") or ""
        if not _text_has_watch_keyword(scan_blob, keyword):
            continue
        if _specific_watchlist_match_exists(scan_blob, keyword, promotions):
            continue
        label = promo.get("label") or keyword
        if label.lower() in existing_blob or keyword.lower() in existing_blob:
            continue
        priority = (promo.get("priority") or "high").lower()
        matched_text = _visible_text_match(visible_text, keyword)
        if not matched_text or not _watchlist_match_has_context(keyword, matched_text, item_context):
            continue
        lead = {
            "label": f"Watchlist match: {label}",
            "why": (promo.get("notes") or f"Matched your scanner watchlist keyword: {keyword}.")[:260],
            "search_phrase": promo.get("search_phrase") or keyword,
            "confidence": "high" if promo.get("comp_link_allowed") and priority == "high" else "medium",
            "location": "Detected text/OCR match; inspect the item in the photo.",
            "evidence": f"Detected text included: {matched_text or keyword}.",
        }
        if priority in {"high", "medium"} and len(focus_first) < 5:
            focus_first.insert(0, lead)
        elif len(maybe_check) < 5:
            maybe_check.append(lead)
        else:
            break
        existing_blob += " " + label.lower() + " " + keyword.lower()

    triage["focus_first"] = focus_first[:5]
    triage["maybe_check"] = maybe_check[:5]
    triage["probably_skip"] = probably_skip[:5]
    return triage


def _save_scanner_watchlist_entries(entries: list, replace_existing: bool = False):
    if replace_existing:
        ScannerWatchKeyword.query.delete()
        db.session.flush()

    imported = 0
    updated = 0
    for entry in entries:
        keyword = (entry.get("keyword") or "").strip()
        if not keyword:
            continue
        existing = (
            ScannerWatchKeyword.query
            .filter(func.lower(ScannerWatchKeyword.keyword) == keyword.lower())
            .first()
        )
        if existing:
            existing.label = entry.get("label") or keyword
            existing.search_phrase = entry.get("search_phrase") or keyword
            existing.category = entry.get("category")
            existing.priority = entry.get("priority") or "high"
            existing.comp_link_allowed = bool(entry.get("comp_link_allowed", True))
            existing.notes = entry.get("notes")
            existing.source = entry.get("source") or existing.source or "scanner-import"
            existing.active = True
            updated += 1
        else:
            db.session.add(ScannerWatchKeyword(
                keyword=keyword,
                label=entry.get("label") or keyword,
                search_phrase=entry.get("search_phrase") or keyword,
                category=entry.get("category"),
                priority=entry.get("priority") or "high",
                comp_link_allowed=bool(entry.get("comp_link_allowed", True)),
                notes=entry.get("notes"),
                source=entry.get("source") or "scanner-import",
                active=True,
            ))
            imported += 1

    db.session.commit()
    return imported, updated


def _sqlite_column_exists(table_name: str, column_name: str) -> bool:
    try:
        rows = db.session.execute(text(f"PRAGMA table_info({table_name})")).fetchall()
        cols = [r[1] for r in rows]  # second field is name
        return column_name in cols
    except Exception:
        return False

def _norm_title(s: str) -> str:
    s = (s or "").lower().strip()
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _similar(a: str, b: str) -> float:
    # cheap fuzzy similarity
    return SequenceMatcher(None, a, b).ratio()

def _parse_ebay_start_date(s: str):
    if not s:
        return None
    parts = s.strip().split()
    if len(parts) >= 3 and parts[-1].isalpha() and len(parts[-1]) in (3, 4):
        s = " ".join(parts[:-1])
    dt = datetime.strptime(s, "%b-%d-%y %H:%M:%S")
    return dt.date()


def _parse_ebay_date(s: str):
    if not s:
        return None
    v = s.strip()
    if not v:
        return None

    # Active listing exports include time + timezone. Order exports are usually date-only.
    for fmt in ("%b %d, %Y", "%b-%d-%Y %H:%M:%S", "%b-%d-%y %H:%M:%S", "%b-%d-%y", "%Y-%m-%d"):
        candidate = v
        if fmt == "%b-%d-%y %H:%M:%S":
            parts = candidate.split()
            if len(parts) >= 3 and parts[-1].isalpha() and len(parts[-1]) in (2, 3, 4):
                candidate = " ".join(parts[:-1])
        try:
            return datetime.strptime(candidate, fmt).date()
        except ValueError:
            continue
    return None


def _detect_ebay_csv(raw: str):
    """
    eBay exports are not consistent: the orders report can include blank preamble rows
    before the actual header. Return (kind, headers, data_rows), where kind is
    "active", "orders", "expenses", or "transactions".
    """
    parsed = list(csv.reader(io.StringIO(raw)))
    for idx, row in enumerate(parsed):
        headers = [(h or "").strip().lstrip("\ufeff") for h in row]
        header_set = set(headers)
        if {"Item number", "Title", "Current price"}.issubset(header_set):
            return "active", headers, parsed[idx + 1:]
        if {"Order Number", "Item Number", "Item Title", "Sold For", "Sale Date"}.issubset(header_set):
            return "orders", headers, parsed[idx + 1:]
        if {"Expense date", "Expense grouping", "Expense category", "Expense type", "Order number", "Net expense"}.issubset(header_set):
            return "expenses", headers, parsed[idx + 1:]
        if {"Transaction creation date", "Type", "Order number", "Net amount", "Item title", "Gross transaction amount"}.issubset(header_set):
            return "transactions", headers, parsed[idx + 1:]
    return None, [], []


def _iter_ebay_rows(raw: str):
    kind, headers, data_rows = _detect_ebay_csv(raw)
    if not kind:
        return None, []

    rows = []
    for row in data_rows:
        if not any((cell or "").strip() for cell in row):
            continue
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        rows.append({headers[i]: padded[i] if i < len(padded) else "" for i in range(len(headers))})
    return kind, rows


def _append_note_tag(item, tag: str):
    tag = (tag or "").strip()
    if not tag:
        return
    note = (item.notes or "").strip()
    if tag not in note:
        item.notes = note + ("\n" if note else "") + tag


def _apply_note_financial_tags(item: Item):
    notes = item.notes or ""
    changed = False
    tag_fields = [
        ("shipping", ("eBayActualShipping",)),
        ("ebay_fee", ("eBayFee", "eBayCollectedCharges")),
        ("ad_fee", ("eBayAdFee",)),
        ("refund_amount", ("eBayRefund",)),
    ]
    for field, tags in tag_fields:
        if getattr(item, field) is not None:
            continue
        for tag in tags:
            match = re.search(rf"(?:^|\n){re.escape(tag)}:([0-9.,-]+)", notes)
            if not match:
                continue
            value = parse_float(match.group(1))
            if value is not None:
                setattr(item, field, abs(value))
                changed = True
                break
    return changed


def _sold_review_expr():
    return (
        (Item.sold.is_(True)) &
        (Item.sold_confirmed.is_(False)) &
        (Item.canceled.is_(False)) &
        (Item.returned.is_(False))
    )


def _pending_shipping_expr():
    return (
        (Item.sold.is_(True)) &
        (Item.sold_confirmed.is_(True)) &
        (Item.pending_shipping.is_(True)) &
        (Item.canceled.is_(False)) &
        (Item.returned.is_(False))
    )


def _shipped_sold_expr():
    return (
        (Item.sold.is_(True)) &
        (Item.sold_confirmed.is_(True)) &
        (Item.pending_shipping.is_(False)) &
        (Item.canceled.is_(False)) &
        (Item.returned.is_(False))
    )


def _returned_expr():
    return (Item.returned.is_(True)) & (Item.canceled.is_(False))


def _is_canceled_order_row(row: dict) -> bool:
    cancel_words = ("cancel", "cancelled", "canceled")
    status_words = ("status", "cancel", "refund")

    for key, value in (row or {}).items():
        key_text = (key or "").strip().lower()
        value_text = (value or "").strip().lower()
        if not value_text:
            continue

        if any(word in key_text for word in status_words):
            if any(phrase in value_text for phrase in ("not cancel", "not cancelled", "not canceled", "no cancel")):
                continue
            if any(word in value_text for word in cancel_words):
                return True
            if "cancel" in key_text and value_text not in {"no", "n", "false", "0", "--"}:
                return True

    return False


def _expense_bucket(row: dict):
    grouping = (row.get("Expense grouping") or "").strip().lower()
    category = (row.get("Expense category") or "").strip().lower()
    expense_type = (row.get("Expense type") or "").strip().lower()
    if "shipping" in grouping or "label" in category or "shipping label" in expense_type:
        return "shipping"
    if "ad fee" in category or "promoted listings" in expense_type:
        return "ad_fee"
    if "transaction fee" in category or "final value fee" in expense_type:
        return "ebay_fee"
    return None


_EBAY_TRANSACTION_FEE_COLUMNS = [
    "Final Value Fee - fixed",
    "Final Value Fee - variable",
    "Regulatory operating fee",
    'Very high "item not as described" fee',
    "Below standard performance fee",
    "International fee",
    "Charity donation",
    "Deposit processing fee",
]


def _transaction_fee_delta(row: dict) -> float:
    """
    Transaction reports show fees as negative numbers and fee credits as positive
    numbers. Store a net positive eBay fee cost for the item.
    """
    fee_cost = 0.0
    for column in _EBAY_TRANSACTION_FEE_COLUMNS:
        amount = parse_float(row.get(column))
        if amount is None:
            continue
        if amount < 0:
            fee_cost += abs(amount)
        elif amount > 0:
            fee_cost -= amount
    return fee_cost


def _transaction_rows_by_order(rows: list[dict]) -> dict:
    grouped = {}

    for row in rows:
        order_number = (row.get("Order number") or row.get("Legacy order ID") or "").strip()
        if not order_number or order_number == "--":
            continue

        transaction_type = (row.get("Type") or "").strip().lower()
        if transaction_type in {"hold", "payout", "transfer", "secondary payout", "reserve", "withheld tax"}:
            continue

        values = grouped.setdefault(order_number, {
            "order_number": order_number,
            "platform": "eBay",
            "price": None,
            "buyer_paid_amount": None,
            "sale_date": None,
            "date_shipped": None,
            "date_returned": None,
            "ebay_item_number": "",
            "ebay_item_url": "",
            "custom_sku": "",
            "title": "",
            "shipping_cost": 0.0,
            "ebay_fee": 0.0,
            "ad_fee": 0.0,
            "refund_amount": 0.0,
            "returned": False,
            "return_reference_id": "",
            "tracking_number": "",
            "transaction_types": set(),
        })
        values["transaction_types"].add(transaction_type or "unknown")

        transaction_date = _parse_ebay_date(row.get("Transaction creation date") or "")
        item_id = (row.get("Item ID") or "").strip()
        title = (row.get("Item title") or "").strip()
        custom_sku = (row.get("Custom label") or "").strip()
        reference_id = (row.get("Reference ID") or "").strip()
        description = (row.get("Description") or "").strip()

        if item_id and item_id != "--" and not values["ebay_item_number"]:
            values["ebay_item_number"] = item_id
            values["ebay_item_url"] = _ebay_item_url(item_id)
        if title and title != "--" and not values["title"]:
            values["title"] = title
        if custom_sku and not values["custom_sku"] and custom_sku != "--":
            values["custom_sku"] = custom_sku

        values["ebay_fee"] += _transaction_fee_delta(row)

        if transaction_type == "order":
            item_subtotal = parse_float(row.get("Item subtotal"))
            shipping_paid = parse_float(row.get("Shipping and handling")) or 0.0
            gross_amount = parse_float(row.get("Gross transaction amount"))
            if item_subtotal is not None:
                values["price"] = item_subtotal
                values["buyer_paid_amount"] = item_subtotal + shipping_paid
            elif gross_amount is not None:
                values["price"] = abs(gross_amount)
                values["buyer_paid_amount"] = abs(gross_amount)
            values["sale_date"] = values["sale_date"] or transaction_date
        elif transaction_type == "shipping label":
            amount = parse_float(row.get("Net amount")) or parse_float(row.get("Gross transaction amount"))
            if amount is not None:
                values["shipping_cost"] += abs(amount)
            values["date_shipped"] = values["date_shipped"] or transaction_date
            if reference_id.lower().startswith("tracking no."):
                values["tracking_number"] = reference_id.split("Tracking no.", 1)[-1].strip()
        elif transaction_type == "other fee":
            amount = parse_float(row.get("Net amount")) or parse_float(row.get("Gross transaction amount"))
            if amount is None:
                continue
            if "promoted listings" in description.lower():
                values["ad_fee"] += abs(amount)
            else:
                values["ebay_fee"] += abs(amount)
        elif transaction_type in {"refund", "claim", "payment dispute", "adjustment"}:
            gross_amount = parse_float(row.get("Gross transaction amount"))
            net_amount = parse_float(row.get("Net amount"))
            refund_amount = gross_amount if gross_amount is not None else net_amount
            if refund_amount is not None and refund_amount < 0:
                values["refund_amount"] += abs(refund_amount)
                values["returned"] = True
                values["date_returned"] = values["date_returned"] or transaction_date
                if reference_id and not values["return_reference_id"]:
                    values["return_reference_id"] = reference_id

    for values in grouped.values():
        values["ebay_fee"] = max(0.0, values["ebay_fee"])
        for key in ("shipping_cost", "ebay_fee", "ad_fee", "refund_amount"):
            values[key] = round(values[key], 2)
            if values[key] == 0:
                values[key] = None
        values["transaction_types"] = ", ".join(sorted(values["transaction_types"]))

    return grouped


def _safe_return_url(value: str):
    value = (value or "").strip()
    if value.startswith("/") and not value.startswith("//"):
        return value
    return url_for("index")


def _copy_missing_item_fields(target: Item, source: Item):
    fields = [
        "category",
        "sub_category",
        "platform",
        "source_location",
        "barcode",
        "ebay_item_number",
        "ebay_order_number",
        "ebay_custom_label",
        "ebay_item_url",
        "ebay_category",
        "ebay_condition",
        "cog",
        "sale_price",
        "ad_fee",
        "ebay_fee",
        "shipping",
        "buyer_paid_amount",
        "refund_amount",
        "returned",
        "return_reference_id",
        "date_returned",
        "date_listed",
        "date_sold",
        "date_shipped",
        "tracking_number",
    ]

    copied = []
    for field in fields:
        if getattr(target, field) in (None, "") and getattr(source, field) not in (None, ""):
            setattr(target, field, getattr(source, field))
            copied.append(field)

    target.sold = bool(target.sold or source.sold)
    target.sold_confirmed = bool(target.sold_confirmed or source.sold_confirmed)
    target.pending_shipping = bool(target.pending_shipping or source.pending_shipping)
    if target.tracking_number or target.date_shipped:
        target.pending_shipping = False
    return copied


def _merge_notes(target: Item, source: Item):
    pieces = []
    existing = (target.notes or "").strip()
    if existing:
        pieces.append(existing)

    source_note = (source.notes or "").strip()
    merge_header = f"Merged from SKU {source.sku}: {source.item_name}"
    if source_note:
        pieces.append(f"{merge_header}\n{source_note}")
    else:
        pieces.append(merge_header)

    target.notes = "\n\n---\n".join(pieces)


def _set_if_missing(item: Item, field: str, value):
    if value in (None, ""):
        return False
    if getattr(item, field) in (None, ""):
        setattr(item, field, value)
        return True
    return False


def _import_value_display(value):
    if value in (None, ""):
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, float):
        return f"${value:.2f}"
    return str(value)


def _preview_fill_if_missing(item: Item, field: str, label: str, value):
    if value in (None, ""):
        return None
    current = getattr(item, field)
    if current in (None, ""):
        return f"{label}: blank → {_import_value_display(value)}"
    return None


def _values_differ(current, incoming):
    if incoming in (None, ""):
        return False
    if current in (None, ""):
        return False
    if isinstance(current, float) or isinstance(incoming, float):
        try:
            return round(float(current), 2) != round(float(incoming), 2)
        except (TypeError, ValueError):
            return str(current) != str(incoming)
    return current != incoming


def _preview_update_if_changed(item: Item, field: str, label: str, value):
    current = getattr(item, field)
    if not _values_differ(current, value):
        return None
    return f"{label}: {_import_value_display(current)} -> {_import_value_display(value)}"


def _set_if_changed(item: Item, field: str, value):
    if not _values_differ(getattr(item, field), value):
        return False
    setattr(item, field, value)
    return True


def _import_change_preview(item: Item, report_kind: str, values: dict):
    if item is None:
        return ["Create new item"] if report_kind not in {"expenses", "transactions"} else []

    changes = []
    if report_kind == "transactions":
        for field, label, key in [
            ("shipping", "Actual shipping", "shipping_cost"),
            ("ebay_fee", "eBay fee", "ebay_fee"),
            ("ad_fee", "Ad fee", "ad_fee"),
            ("sale_price", "Sale price", "price"),
            ("buyer_paid_amount", "Buyer paid before tax", "buyer_paid_amount"),
            ("refund_amount", "Refund amount", "refund_amount"),
            ("date_returned", "Return date", "date_returned"),
            ("return_reference_id", "Return reference", "return_reference_id"),
            ("tracking_number", "Tracking #", "tracking_number"),
            ("date_shipped", "Date shipped", "date_shipped"),
            ("ebay_order_number", "eBay order #", "order_number"),
            ("ebay_item_number", "eBay item #", "ebay_item_number"),
        ]:
            change = _preview_fill_if_missing(item, field, label, values.get(key))
            if change:
                changes.append(change)
            change = _preview_update_if_changed(item, field, label, values.get(key))
            if change:
                changes.append(change)
        if values.get("returned") and not item.returned:
            changes.append("Mark refunded/returned")
        if not item.sold:
            changes.append("Mark sold -> Sold Review")
        return changes

    if report_kind == "expenses":
        change = _preview_fill_if_missing(item, "shipping", "Actual shipping", values.get("shipping_cost"))
        if change:
            changes.append(change)
        change = _preview_fill_if_missing(item, "ebay_fee", "eBay fee", values.get("ebay_fee"))
        if change:
            changes.append(change)
        change = _preview_fill_if_missing(item, "ad_fee", "Ad fee", values.get("ad_fee"))
        if change:
            changes.append(change)
        if not item.sold:
            changes.append("Mark sold → Sold Review")
        return changes

    if report_kind == "orders":
        if values.get("is_canceled_order"):
            if item.sold or item.sold_confirmed or not item.canceled:
                changes.append("Mark canceled, not sold")
            change = _preview_fill_if_missing(item, "ebay_order_number", "eBay order #", values.get("order_number"))
            if change:
                changes.append(change)
            return changes

        if not item.sold:
            changes.append("Mark sold → Sold Review")
            if item.canceled:
                changes.append("Clear canceled flag")
            if values.get("tracking_number") or values.get("date_shipped"):
                if item.pending_shipping:
                    changes.append("Mark shipped / clear Pending Shipping")
                change = _preview_fill_if_missing(item, "tracking_number", "Tracking #", values.get("tracking_number"))
                if change:
                    changes.append(change)
                change = _preview_fill_if_missing(item, "date_shipped", "Date shipped", values.get("date_shipped"))
                if change:
                    changes.append(change)
            for field, label, key in [
                ("platform", "Platform", "platform"),
                ("sale_price", "Sale price", "price"),
                ("ebay_item_number", "eBay item #", "ebay_item_number"),
                ("ebay_item_url", "eBay item URL", "ebay_item_url"),
                ("ebay_custom_label", "Custom label", "custom_sku"),
                ("date_sold", "Date sold", "sale_date"),
                ("buyer_paid_amount", "Buyer paid before tax", "buyer_paid_amount"),
                ("ebay_order_number", "eBay order #", "order_number"),
        ]:
                change = _preview_fill_if_missing(item, field, label, values.get(key))
                if change:
                    changes.append(change)
        return changes

    for field, label, key in [
        ("platform", "Platform", "platform"),
        ("sale_price", "Sale price", "price"),
        ("date_listed", "Date listed", "date_listed"),
        ("ebay_item_number", "eBay item #", "ebay_item_number"),
        ("ebay_item_url", "eBay item URL", "ebay_item_url"),
        ("ebay_custom_label", "Custom label", "custom_sku"),
        ("ebay_category", "eBay category", "category"),
        ("category", "Category", "category"),
        ("ebay_condition", "Condition", "condition"),
    ]:
        change = _preview_fill_if_missing(item, field, label, values.get(key))
        if change:
            changes.append(change)
    price_change = _preview_update_if_changed(item, "sale_price", "Sale price", values.get("price"))
    if price_change:
        changes.append(price_change)
    item_number_change = _preview_update_if_changed(item, "ebay_item_number", "eBay item #", values.get("ebay_item_number"))
    if item_number_change:
        changes.append(item_number_change)
    url_change = _preview_update_if_changed(item, "ebay_item_url", "eBay item URL", values.get("ebay_item_url"))
    if url_change:
        changes.append(url_change)
    if item.canceled:
        changes.append("Clear canceled flag")
    return changes


def _sqlite_add_column(table_name: str, column_name: str, column_type_sql: str):
    db.session.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type_sql}"))
    db.session.commit()


def create_app():
    app = Flask(__name__)
    app.secret_key = os.environ.get("FLASK_SECRET_KEY")
    if not app.secret_key:
        raise RuntimeError("FLASK_SECRET_KEY must be set.")

    app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
        "SQLALCHEMY_DATABASE_URI",
        "sqlite:///ebay_tracker.db"
    )
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("MAX_UPLOAD_BYTES", 20 * 1024 * 1024))
    app.config["MAX_PHOTOS_PER_ITEM"] = int(os.environ.get("MAX_PHOTOS_PER_ITEM", 12))

    default_uploads_dir = Path("/data/uploads/items")
    upload_folder = os.environ.get("UPLOAD_FOLDER", str(default_uploads_dir))
    app.config["UPLOAD_FOLDER"] = upload_folder
    Path(app.config["UPLOAD_FOLDER"]).mkdir(parents=True, exist_ok=True)

    # Defaults for estimator (can be overridden by env vars)
    app.config["EST_EBAY_FEE_PCT"] = float(os.environ.get("EST_EBAY_FEE_PCT", "13.25"))  # %
    app.config["EST_EBAY_FIXED_FEE"] = float(os.environ.get("EST_EBAY_FIXED_FEE", "0.30"))  # $
    app.config["EST_AD_FEE_PCT"] = float(os.environ.get("EST_AD_FEE_PCT", "0"))  # %
    app.config["EST_SHIP_SMALL"] = float(os.environ.get("EST_SHIP_SMALL", "6.50"))
    app.config["EST_SHIP_MED"] = float(os.environ.get("EST_SHIP_MED", "9.50"))
    app.config["EST_SHIP_LARGE"] = float(os.environ.get("EST_SHIP_LARGE", "14.50"))

    db.init_app(app)

    with app.app_context():
        db.create_all()

        # Lightweight “migration” for SQLite for new columns
        if app.config["SQLALCHEMY_DATABASE_URI"].startswith("sqlite"):
            if not _sqlite_column_exists("items", "source_location"):
                _sqlite_add_column("items", "source_location", "VARCHAR(120)")
            if not _sqlite_column_exists("items", "barcode"):
                _sqlite_add_column("items", "barcode", "VARCHAR(64)")
            if not _sqlite_column_exists("items", "sold_confirmed"):
                _sqlite_add_column("items", "sold_confirmed", "BOOLEAN DEFAULT 0 NOT NULL")
            if not _sqlite_column_exists("items", "pending_shipping"):
                _sqlite_add_column("items", "pending_shipping", "BOOLEAN DEFAULT 0 NOT NULL")
            if not _sqlite_column_exists("items", "canceled"):
                _sqlite_add_column("items", "canceled", "BOOLEAN DEFAULT 0 NOT NULL")
            if not _sqlite_column_exists("items", "tracking_number"):
                _sqlite_add_column("items", "tracking_number", "VARCHAR(120)")
            if not _sqlite_column_exists("items", "date_shipped"):
                _sqlite_add_column("items", "date_shipped", "DATE")
            if not _sqlite_column_exists("items", "ebay_item_number"):
                _sqlite_add_column("items", "ebay_item_number", "VARCHAR(32)")
            if not _sqlite_column_exists("items", "ebay_order_number"):
                _sqlite_add_column("items", "ebay_order_number", "VARCHAR(64)")
            if not _sqlite_column_exists("items", "ebay_custom_label"):
                _sqlite_add_column("items", "ebay_custom_label", "VARCHAR(120)")
            if not _sqlite_column_exists("items", "ebay_item_url"):
                _sqlite_add_column("items", "ebay_item_url", "VARCHAR(600)")
            if not _sqlite_column_exists("items", "ebay_category"):
                _sqlite_add_column("items", "ebay_category", "VARCHAR(160)")
            if not _sqlite_column_exists("items", "ebay_condition"):
                _sqlite_add_column("items", "ebay_condition", "VARCHAR(120)")
            if not _sqlite_column_exists("items", "refund_amount"):
                _sqlite_add_column("items", "refund_amount", "FLOAT")
            if not _sqlite_column_exists("items", "returned"):
                _sqlite_add_column("items", "returned", "BOOLEAN DEFAULT 0 NOT NULL")
            if not _sqlite_column_exists("items", "date_returned"):
                _sqlite_add_column("items", "date_returned", "DATE")
            if not _sqlite_column_exists("items", "return_reference_id"):
                _sqlite_add_column("items", "return_reference_id", "VARCHAR(120)")
            if not _sqlite_column_exists("items", "direct_price_override"):
                _sqlite_add_column("items", "direct_price_override", "FLOAT")

    @app.context_processor
    def inject_estimator_defaults():
        # available in all templates
        csrf_token = session.get("_csrf_token")
        if not csrf_token:
            csrf_token = secrets.token_urlsafe(32)
            session["_csrf_token"] = csrf_token

        return dict(
            csrf_token=csrf_token,
            est_defaults={
                "ebay_fee_pct": app.config["EST_EBAY_FEE_PCT"],
                "ebay_fixed_fee": app.config["EST_EBAY_FIXED_FEE"],
                "ad_fee_pct": app.config["EST_AD_FEE_PCT"],
                "ship_small": app.config["EST_SHIP_SMALL"],
                "ship_med": app.config["EST_SHIP_MED"],
                "ship_large": app.config["EST_SHIP_LARGE"],
            }
        )

    @app.before_request
    def protect_post_requests():
        if request.method != "POST":
            return

        expected = session.get("_csrf_token", "")
        supplied = request.form.get("_csrf_token", "")
        if not expected or not supplied or not hmac.compare_digest(expected, supplied):
            abort(400, description="Invalid or missing CSRF token.")
    
    
    # -----------------------------
    # Auth config
    # -----------------------------
    AUTH_MODE = (os.environ.get("AUTH_MODE", "off") or "off").lower()

    BASIC_USER = os.environ.get("BASIC_AUTH_USER", "")
    BASIC_PASS = os.environ.get("BASIC_AUTH_PASS", "")
    BASIC_PASS_HASH = generate_password_hash(BASIC_PASS) if BASIC_PASS else ""

    if AUTH_MODE not in {"off", "basic", "oidc"}:
        raise RuntimeError("AUTH_MODE must be one of: off, basic, oidc.")
    if AUTH_MODE == "basic" and (not BASIC_USER or not BASIC_PASS):
        raise RuntimeError("BASIC_AUTH_USER and BASIC_AUTH_PASS must be set when AUTH_MODE=basic.")
    if AUTH_MODE == "oidc":
        raise RuntimeError("OIDC authentication is not implemented; use AUTH_MODE=basic or off.")

    basic_auth = HTTPBasicAuth()

    @basic_auth.verify_password
    def verify_password(username, password):
        if AUTH_MODE != "basic":
            return False
        if not BASIC_USER or not BASIC_PASS_HASH:
            return False
        return username == BASIC_USER and check_password_hash(BASIC_PASS_HASH, password or "")

    def auth_required(view_func):
        @wraps(view_func)
        def wrapper(*args, **kwargs):
            if AUTH_MODE == "off":
                return view_func(*args, **kwargs)

            if AUTH_MODE == "basic":
                return basic_auth.login_required(view_func)(*args, **kwargs)

            if AUTH_MODE == "oidc":
                if not current_user.is_authenticated:
                    return redirect(url_for("login", next=request.path))
                return view_func(*args, **kwargs)

            return ("Auth misconfigured", 500)

        return wrapper

    @app.get("/export/items.csv")
    @auth_required
    def export_items_csv():
        # Pull all items (you can add filters later)
        items = Item.query.order_by(Item.sku.asc()).all()

        # Build CSV in-memory
        output = io.StringIO()
        w = csv.writer(output)

        # Header row
        w.writerow([
            "sku",
            "item_name",
            "category",
            "sub_category",
            "platform",
            "ebay_item_number",
            "ebay_order_number",
            "ebay_custom_label",
            "ebay_item_url",
            "ebay_category",
            "ebay_condition",
            "barcode",
            "source_location",
            "cog",
            "sale_price",
            "buyer_paid_amount",
            "shipping",
            "ad_fee",
            "ebay_fee",
            "refund_amount",
            "sold",
            "sold_confirmed",
            "pending_shipping",
            "canceled",
            "returned",
            "date_listed",
            "date_sold",
            "date_returned",
            "date_shipped",
            "tracking_number",
            "return_reference_id",
            "notes",
            "image_filenames",
        ])

        for it in items:
            # If you want all image filenames in one column
            try:
                image_names = ";".join([img.filename for img in (it.images or [])])
            except Exception:
                image_names = ""

            w.writerow([
                it.sku,
                it.item_name or "",
                it.category or "",
                it.sub_category or "",
                it.platform or "",
                it.ebay_item_number or "",
                it.ebay_order_number or "",
                it.ebay_custom_label or "",
                it.ebay_item_url or "",
                it.ebay_category or "",
                it.ebay_condition or "",
                it.barcode or "",
                it.source_location or "",
                it.cog if it.cog is not None else "",
                it.sale_price if it.sale_price is not None else "",
                it.buyer_paid_amount if it.buyer_paid_amount is not None else "",
                it.shipping if it.shipping is not None else "",
                it.ad_fee if it.ad_fee is not None else "",
                it.ebay_fee if it.ebay_fee is not None else "",
                it.refund_amount if it.refund_amount is not None else "",
                "Y" if getattr(it, "sold", False) else "N",
                "Y" if getattr(it, "sold_confirmed", False) else "N",
                "Y" if getattr(it, "pending_shipping", False) else "N",
                "Y" if getattr(it, "canceled", False) else "N",
                "Y" if getattr(it, "returned", False) else "N",
                it.date_listed.isoformat() if it.date_listed else "",
                it.date_sold.isoformat() if it.date_sold else "",
                it.date_returned.isoformat() if it.date_returned else "",
                it.date_shipped.isoformat() if it.date_shipped else "",
                it.tracking_number or "",
                it.return_reference_id or "",
                (it.notes or "").replace("\r", " ").replace("\n", " ").strip(),
                image_names,
            ])

        csv_data = output.getvalue()
        output.close()

        return Response(
            csv_data,
            mimetype="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=ebay-tracker-items.csv"
            },
        )

    @app.route("/import/ebay", methods=["GET", "POST"])
    @auth_required
    def import_ebay():
        if request.method == "GET":
            return render_template("import_ebay.html", step="upload")

        uploaded_files = [f for f in request.files.getlist("files") if f and f.filename]
        if not uploaded_files:
            single_file = request.files.get("file")
            if single_file and single_file.filename:
                uploaded_files = [single_file]

        if not uploaded_files:
            flash("Please choose one or more eBay CSV files to upload.", "error")
            return redirect(url_for("import_ebay"))

        file_payloads = []
        kind_order = {"active": 0, "orders": 1, "expenses": 2, "transactions": 3}
        for idx, upload in enumerate(uploaded_files):
            try:
                raw = upload.read().decode("utf-8", errors="replace")
            except Exception:
                flash(f"Could not read {upload.filename}. Make sure it is a CSV exported from eBay.", "error")
                return redirect(url_for("import_ebay"))

            report_kind, parsed_rows = _iter_ebay_rows(raw)
            if not report_kind:
                flash(f"{upload.filename} does not look like an eBay active listings, orders, expenses, or transaction CSV.", "error")
                return redirect(url_for("import_ebay"))

            file_payloads.append({
                "name": secure_filename(upload.filename) or upload.filename or f"upload-{idx + 1}.csv",
                "kind": report_kind,
                "raw": raw,
                "sort_order": kind_order.get(report_kind, 99),
                "original_index": idx,
            })

        file_payloads.sort(key=lambda p: (p["sort_order"], p["original_index"]))

        existing = Item.query.all()
        existing_by_sku = {it.sku: it for it in existing}
        existing_norm = [
            (
                it.sku,
                it.item_name or "",
                _norm_title(it.item_name or ""),
                (getattr(it, "ebay_item_number", None) or "").strip(),
                (getattr(it, "ebay_custom_label", None) or "").strip(),
            )
            for it in existing
        ]
        existing_by_order = {
            (getattr(it, "ebay_order_number", None) or "").strip(): it
            for it in existing
            if (getattr(it, "ebay_order_number", None) or "").strip()
        }

        rows = []
        global_idx = 0

        for file_payload in file_payloads:
            report_kind = file_payload["kind"]
            report_name = file_payload["name"]
            _, parsed_rows = _iter_ebay_rows(file_payload["raw"])

            if report_kind == "transactions":
                grouped = _transaction_rows_by_order(parsed_rows)
                for order_number, values in sorted(grouped.items()):
                    item = existing_by_order.get(order_number)
                    best = None
                    best_score = 0.0
                    match_reason = ""
                    title = values.get("title") or f"eBay transaction {order_number}"
                    ebay_item_number = values.get("ebay_item_number") or ""
                    custom_sku = values.get("custom_sku") or ""

                    if item:
                        best = (item.sku, item.item_name)
                        best_score = 1.0
                        match_reason = "Same eBay order number"
                    else:
                        ntitle = _norm_title(title)
                        for (eid, etitle, entitle, e_item_no, e_custom_label) in existing_norm:
                            if ebay_item_number and e_item_no and ebay_item_number == e_item_no:
                                best = (eid, etitle)
                                best_score = 1.0
                                match_reason = "Same eBay item number"
                                break
                            if custom_sku and e_custom_label and custom_sku == e_custom_label:
                                best = (eid, etitle)
                                best_score = 0.98
                                match_reason = "Same eBay custom label"
                                break
                            if not entitle:
                                continue
                            score = _similar(ntitle, entitle)
                            if score > best_score:
                                best_score = score
                                best = (eid, etitle)
                                match_reason = "Similar title"
                        item = existing_by_sku.get(best[0]) if best and best_score >= 0.86 else None

                    flagged = item is not None
                    expected_changes = _import_change_preview(item, "transactions", values)
                    rows.append({
                        "row_idx": global_idx,
                        "file_name": report_name,
                        "report_kind": report_kind,
                        "title": title,
                        "date_display": values.get("sale_date") or values.get("date_returned") or values.get("date_shipped"),
                        "price": values.get("price") or 0.0,
                        "custom_sku": custom_sku,
                        "ebay_item_number": ebay_item_number,
                        "order_number": order_number,
                        "quantity": 1,
                        "shipping_cost": values.get("shipping_cost"),
                        "ebay_fee": values.get("ebay_fee"),
                        "ad_fee": values.get("ad_fee"),
                        "refund_amount": values.get("refund_amount"),
                        "transaction_types": values.get("transaction_types"),
                        "is_canceled_order": False,
                        "flagged": flagged,
                        "best_match_id": item.sku if item else (best[0] if best else None),
                        "best_match_title": item.item_name if item else (best[1] if best else None),
                        "best_score": round(best_score, 3) if best_score else None,
                        "match_reason": match_reason,
                        "expected_changes": expected_changes,
                        "default_action": "update" if flagged and expected_changes else "skip",
                    })
                    global_idx += 1
                continue

            if report_kind == "expenses":
                grouped = {}
                for r in parsed_rows:
                    order_number = (r.get("Order number") or "").strip()
                    if not order_number or order_number == "--":
                        continue
                    bucket = _expense_bucket(r)
                    if not bucket:
                        continue
                    amount = parse_float(r.get("Net expense"))
                    if amount is None:
                        continue
                    grouped.setdefault(order_number, {"shipping": 0.0, "ebay_fee": 0.0, "ad_fee": 0.0})
                    grouped[order_number][bucket] += abs(amount)

                for order_number, totals in sorted(grouped.items()):
                    item = existing_by_order.get(order_number)
                    expected_changes = _import_change_preview(item, "expenses", {
                        "shipping_cost": totals.get("shipping") or None,
                        "ebay_fee": totals.get("ebay_fee") or None,
                        "ad_fee": totals.get("ad_fee") or None,
                    })
                    rows.append({
                        "row_idx": global_idx,
                        "file_name": report_name,
                        "report_kind": report_kind,
                        "order_number": order_number,
                        "shipping_cost": totals.get("shipping") or None,
                        "ebay_fee": totals.get("ebay_fee") or None,
                        "ad_fee": totals.get("ad_fee") or None,
                        "flagged": item is not None,
                        "best_match_id": item.sku if item else None,
                        "best_match_title": item.item_name if item else None,
                        "expected_changes": expected_changes,
                        "default_action": "update" if item and expected_changes else "skip",
                    })
                    global_idx += 1
                continue

            for r in parsed_rows:
                if report_kind == "orders":
                    title = (r.get("Item Title") or "").strip()
                    ebay_item_number = (r.get("Item Number") or "").strip()
                    price = parse_float(r.get("Sold For")) or 0.0
                    date_display = _parse_ebay_date(r.get("Sale Date") or "")
                    custom_sku = (r.get("Custom Label") or "").strip()
                    order_number = (r.get("Order Number") or "").strip()
                    quantity = parse_int(r.get("Quantity")) or 1
                    is_canceled_order = _is_canceled_order_row(r)
                    tracking_number = (r.get("Tracking Number") or "").strip()
                    date_shipped = _parse_ebay_date(r.get("Shipped On Date") or "")
                    buyer_shipping_paid = parse_float(r.get("Shipping And Handling")) or 0.0
                    total_price = parse_float(r.get("Total Price"))
                    ebay_collected_tax = parse_float(r.get("eBay Collected Tax")) or 0.0
                    ebay_collected_charges = parse_float(r.get("eBay Collected Charges")) or 0.0
                    buyer_paid_amount = (
                        total_price - ebay_collected_tax - ebay_collected_charges
                        if total_price is not None
                        else price + buyer_shipping_paid
                    )
                    category = None
                    condition = None
                else:
                    title = (r.get("Title") or "").strip()
                    ebay_item_number = (r.get("Item number") or "").strip()
                    price = parse_float(r.get("Current price")) or parse_float(r.get("Start price")) or 0.0
                    date_display = _parse_ebay_date(r.get("Start date") or "")
                    custom_sku = (r.get("Custom label (SKU)") or "").strip()
                    order_number = ""
                    quantity = parse_int(r.get("Available quantity")) or 1
                    is_canceled_order = False
                    tracking_number = ""
                    date_shipped = None
                    buyer_paid_amount = None
                    category = (r.get("eBay category 1 name") or "").strip() or None
                    condition = (r.get("Condition") or "").strip() or None

                if not title:
                    continue

                ntitle = _norm_title(title)
                best = None
                best_score = 0.0
                match_reason = ""
                for (eid, etitle, entitle, e_item_no, e_custom_label) in existing_norm:
                    if ebay_item_number and e_item_no and ebay_item_number == e_item_no:
                        best = (eid, etitle)
                        best_score = 1.0
                        match_reason = "Same eBay item number"
                        break
                    if custom_sku and e_custom_label and custom_sku == e_custom_label:
                        best = (eid, etitle)
                        best_score = 0.98
                        match_reason = "Same eBay custom label"
                        break
                    if not entitle:
                        continue
                    score = _similar(ntitle, entitle)
                    if score > best_score:
                        best_score = score
                        best = (eid, etitle)
                        match_reason = "Similar title"

                flagged = best is not None and best_score >= 0.86
                matched_item = existing_by_sku.get(best[0]) if flagged and best else None
                values = {
                    "platform": "eBay",
                    "price": price if price else None,
                    "date_listed": date_display if report_kind != "orders" else None,
                    "sale_date": date_display if report_kind == "orders" else None,
                    "ebay_item_number": ebay_item_number,
                    "ebay_item_url": _ebay_item_url(ebay_item_number),
                    "custom_sku": custom_sku,
                    "category": category,
                    "condition": condition,
                    "order_number": order_number,
                    "buyer_paid_amount": buyer_paid_amount,
                    "is_canceled_order": is_canceled_order,
                    "tracking_number": tracking_number,
                    "date_shipped": date_shipped,
                }
                expected_changes = _import_change_preview(matched_item, report_kind, values)
                rows.append({
                    "row_idx": global_idx,
                    "file_name": report_name,
                    "report_kind": report_kind,
                    "title": title,
                    "date_display": date_display,
                    "price": price,
                    "custom_sku": custom_sku,
                    "ebay_item_number": ebay_item_number,
                    "order_number": order_number,
                    "quantity": quantity,
                    "is_canceled_order": is_canceled_order,
                    "flagged": flagged,
                    "best_match_id": best[0] if best else None,
                    "best_match_title": best[1] if best else None,
                    "best_score": round(best_score, 3),
                    "match_reason": match_reason,
                    "expected_changes": expected_changes,
                    "default_action": "update" if flagged and expected_changes else ("skip" if flagged or is_canceled_order else "create"),
                })
                global_idx += 1

        if not rows:
            flash("No usable rows found in the uploaded file(s).", "warning")
            return redirect(url_for("import_ebay"))

        files_summary = [
            {
                "name": p["name"],
                "kind": p["kind"],
                "label": "Orders" if p["kind"] == "orders" else ("Expenses" if p["kind"] == "expenses" else ("Transactions" if p["kind"] == "transactions" else "Active listings")),
            }
            for p in file_payloads
        ]
        return render_template(
            "import_ebay.html",
            step="preview",
            rows=rows,
            raw_files=json.dumps(file_payloads),
            raw_csv=file_payloads[0]["raw"] if len(file_payloads) == 1 else "",
            report_kind=file_payloads[0]["kind"] if len(file_payloads) == 1 else "multi",
            files_summary=files_summary,
        )

        if report_kind == "expenses":
            grouped = {}
            for r in parsed_rows:
                grouping = (r.get("Expense grouping") or "").strip().lower()
                category = (r.get("Expense category") or "").strip().lower()
                expense_type = (r.get("Expense type") or "").strip().lower()
                order_number = (r.get("Order number") or "").strip()
                if not order_number or order_number == "--":
                    continue
                if "shipping" not in grouping and "label" not in category and "shipping label" not in expense_type:
                    continue
                amount = parse_float(r.get("Net expense"))
                if amount is None:
                    continue
                grouped.setdefault(order_number, 0.0)
                grouped[order_number] += amount

            rows = []
            for i, (order_number, shipping_total) in enumerate(sorted(grouped.items())):
                shipping_cost = abs(shipping_total)
                item = existing_by_order.get(order_number)
                expected_changes = _import_change_preview(item, "expenses", {"shipping_cost": shipping_cost})
                rows.append({
                    "row_idx": i,
                    "order_number": order_number,
                    "shipping_cost": shipping_cost,
                    "flagged": item is not None,
                    "best_match_id": item.sku if item else None,
                    "best_match_title": item.item_name if item else None,
                    "expected_changes": expected_changes,
                    "default_action": "update" if item and expected_changes else "skip",
                })

            if not rows:
                flash("No shipping label expenses with order numbers were found in that file.", "warning")
                return redirect(url_for("import_ebay"))

            return render_template("import_ebay.html", step="preview", rows=rows, raw_csv=raw, report_kind=report_kind)

        rows = []
        for i, r in enumerate(parsed_rows):
            if report_kind == "orders":
                title = (r.get("Item Title") or "").strip()
                ebay_item_number = (r.get("Item Number") or "").strip()
                price = parse_float(r.get("Sold For")) or 0.0
                date_display = _parse_ebay_date(r.get("Sale Date") or "")
                custom_sku = (r.get("Custom Label") or "").strip()
                order_number = (r.get("Order Number") or "").strip()
                quantity = parse_int(r.get("Quantity")) or 1
                is_canceled_order = _is_canceled_order_row(r)
                tracking_number = (r.get("Tracking Number") or "").strip()
                date_shipped = _parse_ebay_date(r.get("Shipped On Date") or "")
                buyer_shipping_paid = parse_float(r.get("Shipping And Handling")) or 0.0
                total_price = parse_float(r.get("Total Price"))
                ebay_collected_tax = parse_float(r.get("eBay Collected Tax")) or 0.0
                ebay_collected_charges = parse_float(r.get("eBay Collected Charges")) or 0.0
                buyer_paid_amount = (
                    total_price - ebay_collected_tax - ebay_collected_charges
                    if total_price is not None
                    else price + buyer_shipping_paid
                )
                category = None
                condition = None
            else:
                title = (r.get("Title") or "").strip()
                ebay_item_number = (r.get("Item number") or "").strip()
                price = parse_float(r.get("Current price")) or parse_float(r.get("Start price")) or 0.0
                date_display = _parse_ebay_date(r.get("Start date") or "")
                custom_sku = (r.get("Custom label (SKU)") or "").strip()
                order_number = ""
                quantity = parse_int(r.get("Available quantity")) or 1
                is_canceled_order = False
                tracking_number = ""
                date_shipped = None
                buyer_paid_amount = None
                category = (r.get("eBay category 1 name") or "").strip() or None
                condition = (r.get("Condition") or "").strip() or None

            if not title:
                continue

            ntitle = _norm_title(title)
            best = None
            best_score = 0.0
            match_reason = ""
            for (eid, etitle, entitle, e_item_no, e_custom_label) in existing_norm:
                if ebay_item_number and e_item_no and ebay_item_number == e_item_no:
                    best = (eid, etitle)
                    best_score = 1.0
                    match_reason = "Same eBay item number"
                    break
                if custom_sku and e_custom_label and custom_sku == e_custom_label:
                    best = (eid, etitle)
                    best_score = 0.98
                    match_reason = "Same eBay custom label"
                    break
                if not entitle:
                    continue
                score = _similar(ntitle, entitle)
                if score > best_score:
                    best_score = score
                    best = (eid, etitle)
                    match_reason = "Similar title"

            flagged = best is not None and best_score >= 0.86
            matched_item = existing_by_sku.get(best[0]) if flagged and best else None
            values = {
                "platform": "eBay",
                "price": price if price else None,
                "date_listed": date_display if report_kind != "orders" else None,
                "sale_date": date_display if report_kind == "orders" else None,
                "ebay_item_number": ebay_item_number,
                "ebay_item_url": _ebay_item_url(ebay_item_number),
                "custom_sku": custom_sku,
                "category": category,
                "condition": condition,
                "order_number": order_number,
                "buyer_paid_amount": buyer_paid_amount,
                "is_canceled_order": is_canceled_order,
                "tracking_number": tracking_number,
                "date_shipped": date_shipped,
            }
            expected_changes = _import_change_preview(matched_item, report_kind, values)
            rows.append({
                "row_idx": i,
                "title": title,
                "date_display": date_display,
                "price": price,
                "custom_sku": custom_sku,
                "ebay_item_number": ebay_item_number,
                "order_number": order_number,
                "quantity": quantity,
                "is_canceled_order": is_canceled_order,
                "flagged": flagged,
                "best_match_id": best[0] if best else None,
                "best_match_title": best[1] if best else None,
                "best_score": round(best_score, 3),
                "match_reason": match_reason,
                "expected_changes": expected_changes,
                "default_action": "update" if flagged and expected_changes else ("skip" if flagged or is_canceled_order else "create"),
            })

        if not rows:
            flash("No usable rows found in that file.", "warning")
            return redirect(url_for("import_ebay"))

        return render_template("import_ebay.html", step="preview", rows=rows, raw_csv=raw, report_kind=report_kind)


    @app.route("/import/ebay/confirm", methods=["POST"])
    @auth_required
    def import_ebay_confirm():
        raw_files = request.form.get("raw_files", "")
        raw_csv = request.form.get("raw_csv", "")
        file_payloads = []

        if raw_files:
            try:
                file_payloads = json.loads(raw_files)
            except Exception:
                file_payloads = []
        elif raw_csv:
            report_kind, _ = _iter_ebay_rows(raw_csv)
            if report_kind:
                file_payloads = [{"name": "upload.csv", "kind": report_kind, "raw": raw_csv}]

        if not file_payloads:
            flash("Import session expired. Please upload again.", "error")
            return redirect(url_for("import_ebay"))

        created = 0
        updated = 0
        skipped = 0
        saw_orders_or_expenses = False
        action_idx = 0

        for file_payload in file_payloads:
            report_kind = file_payload.get("kind")
            raw = file_payload.get("raw", "")
            detected_kind, parsed_rows = _iter_ebay_rows(raw)
            report_kind = report_kind if report_kind == detected_kind else detected_kind
            if not report_kind:
                continue

            if report_kind in ("orders", "expenses", "transactions"):
                saw_orders_or_expenses = True

            if report_kind == "transactions":
                grouped = _transaction_rows_by_order(parsed_rows)
                for order_number, values in sorted(grouped.items()):
                    decision = request.form.get(f"decision_{action_idx}", "skip")
                    match_id = request.form.get(f"matchid_{action_idx}")
                    action_idx += 1
                    if decision == "skip":
                        skipped += 1
                        continue

                    item = Item.query.get(int(match_id)) if match_id else None
                    if not item:
                        skipped += 1
                        continue

                    changed = False
                    changed = _set_if_missing(item, "platform", "eBay") or changed
                    changed = _set_if_missing(item, "sale_price", values.get("price")) or changed
                    changed = _set_if_changed(item, "sale_price", values.get("price")) or changed
                    changed = _set_if_missing(item, "buyer_paid_amount", values.get("buyer_paid_amount")) or changed
                    changed = _set_if_changed(item, "buyer_paid_amount", values.get("buyer_paid_amount")) or changed
                    changed = _set_if_missing(item, "shipping", values.get("shipping_cost")) or changed
                    changed = _set_if_changed(item, "shipping", values.get("shipping_cost")) or changed
                    changed = _set_if_missing(item, "ebay_fee", values.get("ebay_fee")) or changed
                    changed = _set_if_changed(item, "ebay_fee", values.get("ebay_fee")) or changed
                    changed = _set_if_missing(item, "ad_fee", values.get("ad_fee")) or changed
                    changed = _set_if_changed(item, "ad_fee", values.get("ad_fee")) or changed
                    changed = _set_if_missing(item, "refund_amount", values.get("refund_amount")) or changed
                    changed = _set_if_changed(item, "refund_amount", values.get("refund_amount")) or changed
                    changed = _set_if_missing(item, "date_sold", values.get("sale_date")) or changed
                    changed = _set_if_missing(item, "date_shipped", values.get("date_shipped")) or changed
                    changed = _set_if_missing(item, "tracking_number", values.get("tracking_number")) or changed
                    changed = _set_if_missing(item, "date_returned", values.get("date_returned")) or changed
                    changed = _set_if_missing(item, "return_reference_id", values.get("return_reference_id")) or changed
                    changed = _set_if_missing(item, "ebay_order_number", order_number) or changed
                    changed = _set_if_missing(item, "ebay_item_number", values.get("ebay_item_number")) or changed
                    changed = _set_if_missing(item, "ebay_item_url", values.get("ebay_item_url")) or changed
                    changed = _set_if_missing(item, "ebay_custom_label", values.get("custom_sku")) or changed

                    if values.get("returned") and not item.returned:
                        item.returned = True
                        changed = True
                    if not item.sold:
                        item.sold = True
                        changed = True
                    if item.canceled:
                        item.canceled = False
                        changed = True
                    if item.tracking_number or item.date_shipped:
                        if item.pending_shipping:
                            item.pending_shipping = False
                            changed = True
                    if values.get("returned"):
                        item.pending_shipping = False

                    if changed:
                        _append_note_tag(item, f"eBayTransactionReport:{order_number}")
                        if values.get("shipping_cost") is not None:
                            _append_note_tag(item, f"eBayActualShipping:{values['shipping_cost']:.2f}")
                        if values.get("ebay_fee") is not None:
                            _append_note_tag(item, f"eBayFee:{values['ebay_fee']:.2f}")
                        if values.get("ad_fee") is not None:
                            _append_note_tag(item, f"eBayAdFee:{values['ad_fee']:.2f}")
                        if values.get("refund_amount") is not None:
                            _append_note_tag(item, f"eBayRefund:{values['refund_amount']:.2f}")
                        if values.get("return_reference_id"):
                            _append_note_tag(item, f"eBayReturnReference:{values['return_reference_id']}")
                        updated += 1
                    else:
                        skipped += 1
                continue

            if report_kind == "expenses":
                grouped = {}
                for r in parsed_rows:
                    order_number = (r.get("Order number") or "").strip()
                    if not order_number or order_number == "--":
                        continue
                    bucket = _expense_bucket(r)
                    if not bucket:
                        continue
                    amount = parse_float(r.get("Net expense"))
                    if amount is None:
                        continue
                    grouped.setdefault(order_number, {"shipping": 0.0, "ebay_fee": 0.0, "ad_fee": 0.0})
                    grouped[order_number][bucket] += abs(amount)

                for order_number, totals in sorted(grouped.items()):
                    decision = request.form.get(f"decision_{action_idx}", "skip")
                    match_id = request.form.get(f"matchid_{action_idx}")
                    action_idx += 1
                    if decision == "skip":
                        skipped += 1
                        continue

                    item = Item.query.get(int(match_id)) if match_id else None
                    if not item:
                        skipped += 1
                        continue

                    changed = False
                    shipping_cost = totals.get("shipping") or None
                    ebay_fee = totals.get("ebay_fee") or None
                    ad_fee = totals.get("ad_fee") or None
                    changed = _set_if_missing(item, "shipping", shipping_cost) or changed
                    changed = _set_if_missing(item, "ebay_fee", ebay_fee) or changed
                    changed = _set_if_missing(item, "ad_fee", ad_fee) or changed
                    if not item.sold:
                        item.sold = True
                        changed = True
                    if changed:
                        if shipping_cost is not None:
                            _append_note_tag(item, f"eBayActualShipping:{shipping_cost:.2f}")
                        if ebay_fee is not None:
                            _append_note_tag(item, f"eBayFee:{ebay_fee:.2f}")
                        if ad_fee is not None:
                            _append_note_tag(item, f"eBayAdFee:{ad_fee:.2f}")
                        updated += 1
                    else:
                        skipped += 1
                continue

            for r in parsed_rows:
                if report_kind == "orders":
                    title = (r.get("Item Title") or "").strip()
                    ebay_item_number = (r.get("Item Number") or "").strip()
                    order_number = (r.get("Order Number") or "").strip()
                    custom_sku = (r.get("Custom Label") or "").strip()
                    is_canceled_order = _is_canceled_order_row(r)
                    tracking_number = (r.get("Tracking Number") or "").strip()
                    date_shipped = _parse_ebay_date(r.get("Shipped On Date") or "")
                    price = parse_float(r.get("Sold For")) or 0.0
                    buyer_shipping_paid = parse_float(r.get("Shipping And Handling")) or 0.0
                    total_price = parse_float(r.get("Total Price"))
                    ebay_collected_tax = parse_float(r.get("eBay Collected Tax")) or 0.0
                    ebay_collected_charges = parse_float(r.get("eBay Collected Charges")) or 0.0
                    buyer_paid_amount = (
                        total_price - ebay_collected_tax - ebay_collected_charges
                        if total_price is not None
                        else price + buyer_shipping_paid
                    )
                    sale_date = _parse_ebay_date(r.get("Sale Date") or "")
                    date_listed = None
                    category = None
                    condition = None
                else:
                    title = (r.get("Title") or "").strip()
                    ebay_item_number = (r.get("Item number") or "").strip()
                    order_number = ""
                    custom_sku = (r.get("Custom label (SKU)") or "").strip()
                    is_canceled_order = False
                    tracking_number = ""
                    date_shipped = None
                    price = parse_float(r.get("Current price")) or parse_float(r.get("Start price")) or 0.0
                    buyer_paid_amount = None
                    buyer_shipping_paid = None
                    ebay_collected_tax = None
                    ebay_collected_charges = None
                    sale_date = None
                    date_listed = _parse_ebay_date(r.get("Start date") or "")
                    category = (r.get("eBay category 1 name") or "").strip() or None
                    condition = (r.get("Condition") or "").strip() or None

                if not title:
                    continue

                decision = request.form.get(f"decision_{action_idx}", "skip")
                match_id = request.form.get(f"matchid_{action_idx}")
                action_idx += 1

                if decision == "skip":
                    skipped += 1
                    continue

                if report_kind == "orders" and is_canceled_order and decision == "create":
                    skipped += 1
                    continue

                if decision == "update":
                    item = Item.query.get(int(match_id)) if match_id else None
                    if not item:
                        skipped += 1
                        continue
                else:
                    item = Item(item_name=title)
                    db.session.add(item)
                    created += 1

                changed = decision == "create"

                if decision == "create":
                    item.item_name = title
                    item.platform = "eBay"
                    item.canceled = False
                    if price:
                        item.sale_price = price
                    if date_listed:
                        item.date_listed = date_listed
                    if ebay_item_number:
                        item.ebay_item_number = ebay_item_number
                        item.ebay_item_url = _ebay_item_url(ebay_item_number)
                    if custom_sku:
                        item.ebay_custom_label = custom_sku
                    if category:
                        item.ebay_category = category
                        item.category = category
                    if condition:
                        item.ebay_condition = condition
                else:
                    changed = _set_if_missing(item, "platform", "eBay") or changed
                    if report_kind == "orders":
                        changed = _set_if_missing(item, "sale_price", price if price else None) or changed
                    else:
                        changed = (
                            _set_if_missing(item, "sale_price", price if price else None) or
                            _set_if_changed(item, "sale_price", price if price else None) or
                            changed
                        )
                    changed = _set_if_missing(item, "date_listed", date_listed) or changed
                    if report_kind == "orders":
                        changed = _set_if_missing(item, "ebay_item_number", ebay_item_number) or changed
                    else:
                        changed = (
                            _set_if_missing(item, "ebay_item_number", ebay_item_number) or
                            _set_if_changed(item, "ebay_item_number", ebay_item_number) or
                            changed
                        )
                    if ebay_item_number:
                        if report_kind == "orders":
                            changed = _set_if_missing(item, "ebay_item_url", _ebay_item_url(ebay_item_number)) or changed
                        else:
                            changed = (
                                _set_if_missing(item, "ebay_item_url", _ebay_item_url(ebay_item_number)) or
                                _set_if_changed(item, "ebay_item_url", _ebay_item_url(ebay_item_number)) or
                                changed
                            )
                    changed = _set_if_missing(item, "ebay_custom_label", custom_sku) or changed
                    changed = _set_if_missing(item, "ebay_category", category) or changed
                    changed = _set_if_missing(item, "category", category) or changed
                    changed = _set_if_missing(item, "ebay_condition", condition) or changed

                if report_kind == "orders":
                    if is_canceled_order:
                        if item.sold or item.sold_confirmed or not item.canceled:
                            changed = True
                        item.sold = False
                        item.sold_confirmed = False
                        item.pending_shipping = False
                        item.canceled = True
                        changed = _set_if_missing(item, "ebay_order_number", order_number) or changed
                        if changed:
                            _append_note_tag(item, f"eBayCanceledOrder:{order_number}")
                    elif not item.sold:
                        item.sold = True
                        item.canceled = False
                        changed = True
                    elif item.canceled:
                        item.canceled = False
                        changed = True

                    if not is_canceled_order:
                        changed = _set_if_missing(item, "date_sold", sale_date) or changed
                        changed = _set_if_missing(item, "buyer_paid_amount", buyer_paid_amount) or changed
                        changed = _set_if_missing(item, "ebay_order_number", order_number) or changed
                        review_changed = changed
                        changed = _set_if_missing(item, "tracking_number", tracking_number) or changed
                        changed = _set_if_missing(item, "date_shipped", date_shipped) or changed
                        if (tracking_number or date_shipped) and item.pending_shipping:
                            item.pending_shipping = False
                            changed = True
                        if changed:
                            _append_note_tag(item, f"eBayOrder:{order_number}")
                            _append_note_tag(item, f"eBaySoldFor:{price:.2f}")
                            _append_note_tag(item, f"eBayBuyerShippingPaid:{buyer_shipping_paid:.2f}")
                            _append_note_tag(item, f"eBayCollectedTax:{ebay_collected_tax:.2f}")
                            if ebay_collected_charges:
                                _append_note_tag(item, f"eBayCollectedCharges:{ebay_collected_charges:.2f}")
                else:
                    item.sold = False if item.sold is None else item.sold
                    if item.canceled:
                        item.canceled = False
                        changed = True

                if decision == "update":
                    if changed:
                        _append_note_tag(item, f"eBaySKU:{custom_sku}")
                        updated += 1
                    else:
                        skipped += 1
                elif decision == "create":
                    _append_note_tag(item, f"eBaySKU:{custom_sku}")

        db.session.commit()
        label = "file" if len(file_payloads) == 1 else "files"
        flash(f"Imported eBay {label}. Created: {created}, Updated: {updated}, Skipped: {skipped}.", "success")
        return redirect(url_for("index", status="sold_review" if saw_orders_or_expenses else "all"))

    @app.route("/import/ebay/active", methods=["GET", "POST"])
    @auth_required
    def import_ebay_active():
        if request.method == "GET":
            return render_template("import_ebay_active.html", step="upload")

        # --- step 1: parse upload and show preview ---
        f = request.files.get("file")
        if not f or not f.filename:
            flash("Please choose the eBay CSV file to upload.", "error")
            return redirect(url_for("import_ebay_active"))

        try:
            raw = f.read().decode("utf-8", errors="replace")
        except Exception:
            flash("Could not read that file. Make sure it’s a CSV exported from eBay.", "error")
            return redirect(url_for("import_ebay_active"))

        reader = csv.DictReader(io.StringIO(raw))
        required_cols = ["Title", "Start date", "Current price", "Custom label (SKU)"]

        missing = [c for c in required_cols if c not in (reader.fieldnames or [])]
        if missing:
            flash(f"Missing expected columns: {', '.join(missing)}", "error")
            return redirect(url_for("import_ebay_active"))

        # Pull existing items for matching
        existing = Item.query.all()
        existing_norm = [(it.sku, (it.item_name or ""), _norm_title(it.item_name or "")) for it in existing]

        rows = []
        for i, r in enumerate(reader):
            title = (r.get("Title") or "").strip()
            if not title:
                continue

            start_date_raw = (r.get("Start date") or "").strip()
            price_raw = (r.get("Current price") or "").strip()
            sku_raw = (r.get("Custom label (SKU)") or "").strip()

            # parse price
            try:
                price = float(str(price_raw).replace("$", "").replace(",", "").strip() or 0)
            except Exception:
                price = 0.0

            # parse date -> MM-DD-YYYY
            try:
                date_listed = _parse_ebay_start_date(start_date_raw)
            except Exception:
                date_listed = ""

            ntitle = _norm_title(title)

            # find best match by similarity
            best = None
            best_score = 0.0
            for (eid, etitle, entitle) in existing_norm:
                if not entitle:
                    continue
                score = _similar(ntitle, entitle)
                if score > best_score:
                    best_score = score
                    best = (eid, etitle)

            # threshold for "flag as possible duplicate"
            flagged = best is not None and best_score >= 0.86

            rows.append({
                "row_idx": i,
                "title": title,
                "date_listed": date_listed,
                "price": price,
                "custom_sku": sku_raw,
                "flagged": flagged,
                "best_match_id": best[0] if best else None,
                "best_match_title": best[1] if best else None,
                "best_score": round(best_score, 3),
            })

        if not rows:
            flash("No rows found in that file.", "warning")
            return redirect(url_for("import_ebay_active"))

        # Send preview rows to template (we’ll re-post them as hidden JSON)
        return render_template("import_ebay_active.html", step="preview", rows=rows, raw_csv=raw)


    @app.route("/import/ebay/active/confirm", methods=["POST"])
    @auth_required
    def import_ebay_active_confirm():
        raw_csv = request.form.get("raw_csv", "")
        if not raw_csv:
            flash("Import session expired (missing payload). Please upload again.", "error")
            return redirect(url_for("import_ebay_active"))

        reader = csv.DictReader(io.StringIO(raw_csv))
        # decisions come back as decision_<row_idx>
        created = 0
        updated = 0
        skipped = 0

        for i, r in enumerate(reader):
            title = (r.get("Title") or "").strip()
            if not title:
                continue

            decision = request.form.get(f"decision_{i}", "skip")  # skip | update | create
            if decision == "skip":
                skipped += 1
                continue

            start_date_raw = (r.get("Start date") or "").strip()
            price_raw = (r.get("Current price") or "").strip()
            sku_raw = (r.get("Custom label (SKU)") or "").strip()

            try:
                price = float(str(price_raw).replace("$", "").replace(",", "").strip() or 0)
            except Exception:
                price = 0.0

            try:
                date_listed = _parse_ebay_start_date(start_date_raw)
            except Exception:
                date_listed = ""

            # If updating, we need the matched id from the form
            if decision == "update":
                match_id = request.form.get(f"matchid_{i}")
                if not match_id:
                    skipped += 1
                    continue

                item = Item.query.get(int(match_id))
                if not item:
                    skipped += 1
                    continue

                # Update a few safe fields (don’t touch COG, etc.)
                item.item_name = title
                if price:
                    item.sale_price = price
                if date_listed:
                    item.date_listed = date_listed
                item.canceled = False

                # tuck custom SKU into notes (so you have it even before you add a real column)
                if sku_raw:
                    note = (item.notes or "").strip()
                    tag = f"eBaySKU:{sku_raw}"
                    if tag not in note:
                        item.notes = (note + ("\n" if note else "") + tag)

                updated += 1
                continue

            # Otherwise create new
            item = Item(
                item_name=title,
                sale_price=price if price else None,
                date_listed=date_listed or None,
                sold=False,  # adjust if your model uses boolean, etc.
                canceled=False,
            )

            if sku_raw:
                item.notes = f"eBaySKU:{sku_raw}"

            db.session.add(item)
            created += 1

        db.session.commit()
        flash(f"Import complete. Created: {created}, Updated: {updated}, Skipped: {skipped}.", "success")
        return redirect(url_for("index"))

    @app.get("/tools/scanner")
    @auth_required
    def scanner_tool():
        watchlist_count = ScannerWatchKeyword.query.filter(ScannerWatchKeyword.active.is_(True)).count()
        return render_template("scanner_tool.html", watchlist_count=watchlist_count)

    @app.post("/tools/scanner/watchlist/import")
    @auth_required
    def scanner_watchlist_import():
        file_storage = request.files.get("watchlist_file")
        replace_existing = request.form.get("replace_existing") == "1"
        try:
            entries = _parse_scanner_watchlist_upload(file_storage)
        except ValueError as exc:
            flash(str(exc), "error")
            return redirect(url_for("scanner_tool"))

        if not entries:
            flash("No watchlist keywords were found in that file.", "error")
            return redirect(url_for("scanner_tool"))

        imported, updated = _save_scanner_watchlist_entries(entries, replace_existing=replace_existing)
        flash(
            f"Scanner watchlist imported. Added {imported}, updated {updated}.",
            "success",
        )
        return redirect(url_for("scanner_tool"))

    @app.post("/tools/scanner/quick-create")
    @auth_required
    def scanner_quick_create():
        barcode = request.form.get("barcode", "").strip()
        title = request.form.get("draft_title", "").strip() or barcode
        source_location = request.form.get("source_location", "").strip() or None
        cog = parse_float(request.form.get("cog"))
        sale_price = parse_float(request.form.get("sale_price"))
        shipping = parse_float(request.form.get("shipping"))

        if not title:
            flash("Scan a barcode or type a rough title first.", "error")
            return redirect(url_for("scanner_tool"))

        notes = [
            "ScannerDraft:yes",
            f"ScannerSearch:{title}",
        ]
        if barcode:
            notes.append(f"ScannerBarcode:{barcode}")
        notes.append("Created from fast mobile scanner. Rename/merge after final eBay listing if needed.")

        item = Item(
            item_name=title,
            platform="eBay",
            barcode=barcode or None,
            source_location=source_location,
            cog=cog,
            sale_price=sale_price,
            shipping=shipping,
            sold=False,
            notes="\n".join(notes),
        )
        db.session.add(item)
        db.session.commit()
        flash(f"Quick draft created: SKU #{item.sku}.", "success")
        return redirect(url_for("scanner_tool", created=item.sku))

    @app.post("/tools/scanner/shelf-triage")
    @auth_required
    def scanner_shelf_triage():
        photo = request.files.get("photo")
        if not photo:
            return jsonify({"ok": False, "error": "Choose a shelf photo first."}), 400

        try:
            image_data_url = _prepare_shelf_triage_image(photo)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400

        local_base = (
            os.environ.get("VISION_API_BASE")
            or os.environ.get("LM_STUDIO_URL")
            or ""
        ).strip()
        local_endpoint = _vision_api_endpoint(local_base)
        using_local_vision = bool(local_endpoint)
        api_key = (
            os.environ.get("VISION_API_KEY")
            or os.environ.get("OPENAI_API_KEY")
            or ""
        ).strip()

        if not using_local_vision and not api_key:
            return jsonify({
                "ok": False,
                "error": "Set LM_STUDIO_URL/VISION_API_BASE for local vision, or set OPENAI_API_KEY for hosted vision.",
            }), 400

        watchlist_promotions = _scanner_watchlist_promotions()
        history_promotions = _scanner_history_keyword_promotions()
        watchlist_hint = ""
        if watchlist_promotions:
            watchlist_hint = (
                "The app will compare your response against the user's scanner watchlist after you answer, "
                "so do not invent brands, franchises, or categories from a hidden list. Only name watchlist-like "
                "items when they are visibly readable or visually identifiable in the photo. "
            )
        history_hint = ""
        if history_promotions:
            history_terms = ", ".join(p["keyword"] for p in history_promotions[:35])
            history_hint = (
                "Also prioritize visible terms related to the user's confirmed sold history. "
                f"Sold-history keywords include: {history_terms}. "
            )

        model = (
            os.environ.get("VISION_MODEL")
            or os.environ.get("OPENAI_VISION_MODEL")
            or ("local-vision" if using_local_vision else "gpt-5.5")
        ).strip()
        max_tokens = parse_int(os.environ.get("VISION_MAX_TOKENS")) or 1600
        vision_temperature = parse_float(os.environ.get("VISION_TEMPERATURE"))
        if vision_temperature is None:
            vision_temperature = 0
        json_system_prompt = (
            "You are a JSON API for a reseller shelf-photo triage tool. "
            "Return only one valid JSON object. Do not include markdown, chain-of-thought, analysis, or explanations outside JSON."
        )
        prompt = (
            "/no_think\n"
            "You are helping a fast-moving eBay reseller triage a shelf photo in a store or storage room. "
            "The user is walking up to a table or shelf and wants to know what to physically pick up first. "
            "This is not inventory creation and not a full image description. Identify visible items, brands, "
            "titles, model numbers, UPCs, logos, sealed packaging, recognizable characters, and hidden-value categories "
            "that are worth checking for labels, markings, or sold comps. Be honest about uncertainty and do not invent text "
            "you cannot read. Do not waste slots on obvious trash, papers, clutter, generic cables, rulers, glue, "
            "stationery, loose books, or vague searches like 'gray toy', 'toys', 'generic books', or 'box on shelf'. "
            "Do not return layout/meta entries such as 'Image Analysis', 'Left Shelf', 'Second shelf down', "
            "'Bookshelf', 'TV stand', or any shelf/location name. Every label must be a physical product/object "
            "the reseller can pick up, inspect, or comp. "
            "Balance readable labels with treasure-hunt categories: include unknown figurines, statues, action figures, "
            "dolls, plush, glass, porcelain, ceramic, crystal, brass/metal/wood pieces, vintage toys, sealed items, or "
            "anything with a possible maker's mark/tag/stamp even if you cannot identify the exact brand. For those, use "
            "a useful action phrase like 'inspect bottom for maker mark' or 'check tag/character ID' instead of a vague search. "
            "For every returned item, include a short location like 'left side of table', 'center under blue strap', "
            "or 'top right box', plus evidence explaining what you saw, such as readable text, logo, shape, tag, "
            "model number, or visual clue. If the brand/title is uncertain, say that in evidence. "
            "The visible_text array must contain only text physically visible in the image. Never include prompt words "
            "or internal labels such as Watchlist, Sold History, Label, Search phrase, Confidence, Where, or Evidence. "
            "Always surface visible gaming/fantasy/tabletop collector keywords as leads: Zelda/Legend of Zelda, Nintendo, "
            "Dungeons & Dragons/D&D/DND/AD&D, TSR, Pathfinder, Warhammer, Magic: The Gathering/MTG, Pokemon, Yu-Gi-Oh, "
            "Lord of the Rings/Hobbit, Star Wars, anime/manga, comics, and similar fandom/IP items. These can be valuable "
            "even when they are books/manuals/cards rather than toys. "
            f"{watchlist_hint}"
            f"{history_hint}"
            "Use confidence 'high' only when you can read a specific product title, brand+model, UPC, or exact named item suitable "
            "for an eBay sold search. For partial/generic descriptions like 'Star Wars box', 'green plush toy', 'white ceramic toilet', "
            "or 'small box', use medium or low and phrase it as an inspection lead. "
            "If something is low-value clutter, omit it entirely instead of putting it in probably_skip. "
            "Return only valid JSON with this exact shape: "
            '{"summary":"short practical summary","focus_first":[{"label":"item","why":"why it may be worth checking","location":"where in photo","evidence":"what visible clue supports it","search_phrase":"best eBay sold search phrase","confidence":"low|medium|high"}],'
            '"maybe_check":[{"label":"item","why":"why maybe","location":"where in photo","evidence":"what visible clue supports it","search_phrase":"search phrase","confidence":"low|medium|high"}],'
            '"probably_skip":[{"label":"item","why":"why likely low priority","location":"where in photo","evidence":"what visible clue supports it","search_phrase":"search phrase","confidence":"low|medium|high"}],'
            '"visible_text":["short visible text snippets"]}. '
            "Use focus_first for 'pick up first' items with readable brand/IP/model or strong resale category. "
            "Use maybe_check for 'inspect closer' items that could be valuable if marked, tagged, complete, vintage, or recognizable up close. "
            "Limit focus_first to the best 5 real leads. Limit maybe_check to real maybe leads. Use probably_skip only "
            "for reseller-relevant warnings, not ordinary clutter; it may be empty. Keep each reason short. If unsure, use confidence low."
        )

        headers = {"Content-Type": "application/json"}
        if api_key:
            headers["Authorization"] = f"Bearer {api_key}"

        try:
            if using_local_vision:
                use_response_format = (os.environ.get("VISION_RESPONSE_FORMAT", "json_object") or "").lower() != "off"
                body = {
                    "model": model,
                    "messages": [
                        {"role": "system", "content": json_system_prompt},
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": prompt},
                                {"type": "image_url", "image_url": {"url": image_data_url}},
                            ],
                        },
                    ],
                    "max_tokens": max_tokens,
                    "temperature": vision_temperature,
                    "top_p": 0.1,
                }
                if use_response_format:
                    body["response_format"] = {"type": "json_object"}
                try:
                    payload = _post_vision_json(local_endpoint, body, headers)
                except urllib.error.HTTPError as exc:
                    detail = exc.read().decode("utf-8", errors="replace")[:700]
                    if not use_response_format:
                        raise
                    body.pop("response_format", None)
                    try:
                        payload = _post_vision_json(local_endpoint, body, headers)
                    except urllib.error.HTTPError as retry_exc:
                        retry_detail = retry_exc.read().decode("utf-8", errors="replace")[:700]
                        return jsonify({
                            "ok": False,
                            "error": (
                                f"Vision model failed with response_format ({exc.code}): {detail}\n\n"
                                f"Retry without response_format also failed ({retry_exc.code}): {retry_detail}"
                            ),
                        }), 502
            else:
                body = {
                    "model": model,
                    "input": [{
                        "role": "user",
                        "content": [
                            {"type": "input_text", "text": f"{json_system_prompt}\n\n{prompt}"},
                            {"type": "input_image", "image_url": image_data_url, "detail": "low"},
                        ],
                    }],
                    "max_output_tokens": max_tokens,
                }
                payload = _post_vision_json("https://api.openai.com/v1/responses", body, headers)
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")[:700]
            return jsonify({
                "ok": False,
                "error": f"Vision model could not analyze the photo ({exc.code}). {detail}",
            }), 502
        except Exception as exc:
            return jsonify({
                "ok": False,
                "error": f"Vision photo analysis failed: {exc}",
            }), 502

        text = _extract_chat_completion_text(payload) if using_local_vision else _extract_openai_response_text(payload)
        if not text:
            return jsonify({"ok": False, "error": "Vision model returned no triage text."}), 502

        try:
            triage = _normalize_shelf_triage(_loads_json_object(text))
        except Exception:
            fallback = _parse_shelf_triage_text_fallback(text)
            if fallback:
                triage = _normalize_shelf_triage(fallback)
            else:
                triage = _normalize_shelf_triage({
                    "summary": "The vision model returned text I could not bucket. Try again, or use a model that supports JSON/structured output.",
                    "focus_first": [],
                    "maybe_check": [],
                    "probably_skip": [],
                    "visible_text": [],
                })

        triage = _apply_watchlist_triage_boosts(triage, watchlist_promotions)
        triage = _apply_history_triage_boosts(triage, history_promotions)
        triage = _prioritize_triage_buckets(triage)
        return jsonify({"ok": True, "triage": triage})


    @app.route("/uploads/items/<path:filename>")
    @auth_required
    def uploaded_file(filename):
        return send_from_directory(app.config["UPLOAD_FOLDER"], filename)

    def _public_store_query():
        listed_expr = (
            (Item.ebay_item_number.isnot(None)) &
            (func.trim(Item.ebay_item_number) != "")
        )
        return (
            Item.query
            .filter(Item.sold.is_(False), Item.canceled.is_(False), Item.returned.is_(False))
            .filter(listed_expr)
        )

    def _public_store_department(item):
        text = " ".join([
            item.item_name or "",
            item.category or "",
            item.sub_category or "",
            item.ebay_category or "",
        ]).lower()
        rules = [
            ("Video Games", ("video game", "nintendo", "playstation", "xbox", "wii", "switch", "console", "controller", "gamecube", "sega")),
            ("Board Games & Toys", ("board game", "card game", "party game", "toy", "figure", "figurine", "doll", "plush", "lego", "puzzle", "action figure")),
            ("Trading Cards", ("trading card", "sports card", "pokemon", "mtg", "magic the gathering", "yugioh", "card lot")),
            ("Books", ("book", "books", "novel", "hardcover", "paperback", "manga", "comic", "guide", "manual")),
            ("Movies & Music", ("dvd", "blu-ray", "bluray", "vhs", "cassette", "cd", "vinyl", "record", "album", "movie", "music")),
            ("Clothing", ("shirt", "t-shirt", "tee", "hoodie", "jacket", "pants", "jeans", "hat", "cap", "shoes", "dress", "sweater", "clothing")),
            ("Electronics", ("electronics", "computer", "pc", "laptop", "tablet", "phone", "camera", "scanner", "printer", "receiver", "speaker", "headphone", "gpu", "graphics card")),
            ("Collectibles", ("collectible", "collectibles", "vintage", "memorabilia", "ornament", "coin", "stamp", "rare")),
            ("Home & Decor", ("home", "decor", "kitchen", "glass", "ceramic", "porcelain", "vase", "bowl", "plate", "mug", "lamp", "art")),
            ("Sports & Outdoors", ("sport", "sports", "golf", "baseball", "football", "basketball", "fishing", "camping", "outdoor")),
            ("Tools & Parts", ("tool", "tools", "part", "parts", "hardware", "automotive", "cable", "adapter", "charger")),
        ]
        for label, keywords in rules:
            if any(keyword in text for keyword in keywords):
                return label
        return "Other"

    def _app_setting_float(key, default, min_value=None, max_value=None, env_key=None):
        setting = db.session.get(AppSetting, key)
        raw = setting.value if setting and setting.value not in (None, "") else None
        if raw is None and env_key:
            raw = os.environ.get(env_key)
        if raw is None:
            raw = default
        try:
            value = float(raw)
        except (TypeError, ValueError):
            value = float(default)
        if min_value is not None:
            value = max(float(min_value), value)
        if max_value is not None:
            value = min(float(max_value), value)
        return value

    def _set_app_setting(key, value):
        setting = db.session.get(AppSetting, key)
        if setting is None:
            setting = AppSetting(key=key)
            db.session.add(setting)
        setting.value = f"{value:g}"

    def _public_store_discount_percent():
        return _app_setting_float(
            "store_direct_discount_percent",
            10.0,
            min_value=0.0,
            max_value=95.0,
            env_key="STORE_DIRECT_DISCOUNT_PERCENT",
        )

    def _store_min_profit_warning():
        return _app_setting_float("store_min_profit_warning", 5.0, min_value=0.0, max_value=999999.0)

    def _store_min_margin_warning():
        return _app_setting_float("store_min_margin_warning", 40.0, min_value=0.0, max_value=100.0)

    def _decorate_public_store_item(item, discount_percent):
        item.store_department = _public_store_department(item)
        item.direct_discount_percent = discount_percent
        item.direct_price = None
        item.direct_price_is_override = item.direct_price_override is not None
        item.direct_discount_amount = 0.0
        if item.direct_price_override is not None:
            item.direct_price = round(float(item.direct_price_override), 2)
            if item.sale_price and float(item.sale_price) > 0:
                item.direct_discount_percent = max(0.0, (1 - (item.direct_price / float(item.sale_price))) * 100.0)
            else:
                item.direct_discount_percent = 0.0
        elif item.sale_price is not None and discount_percent > 0:
            item.direct_price = round(float(item.sale_price) * (1 - discount_percent / 100.0), 2)
        if item.sale_price is not None and item.direct_price is not None:
            item.direct_discount_amount = max(0.0, round(float(item.sale_price) - float(item.direct_price), 2))
        return item

    def _public_store_effective_price(item):
        if item.direct_price is not None:
            return float(item.direct_price)
        if item.sale_price is not None:
            return float(item.sale_price)
        return None

    @app.get("/settings/store")
    @auth_required
    def store_settings():
        direct_discount_percent = _public_store_discount_percent()
        min_profit_warning = _store_min_profit_warning()
        min_margin_warning = _store_min_margin_warning()
        store_items = _public_store_query().order_by(Item.date_listed.desc(), Item.sku.desc()).all()
        preview_rows = []
        missing_cog_count = 0
        warning_count = 0
        total_direct_profit = 0.0

        for item in store_items:
            _decorate_public_store_item(item, direct_discount_percent)
            direct_price = item.direct_price if item.direct_price is not None else item.sale_price
            cog = item.cog
            projected_profit = None
            projected_margin = None
            warnings = []

            if cog is None:
                missing_cog_count += 1
                warnings.append("Missing COG")
            elif direct_price is not None:
                projected_profit = float(direct_price) - float(cog)
                projected_margin = (projected_profit / float(direct_price) * 100.0) if float(direct_price) else None
                total_direct_profit += projected_profit
                if projected_profit < min_profit_warning:
                    warnings.append("Low profit")
                if projected_margin is not None and projected_margin < min_margin_warning:
                    warnings.append("Low margin")

            if warnings:
                warning_count += 1

            preview_rows.append(
                {
                    "item": item,
                    "direct_price": direct_price,
                    "projected_profit": projected_profit,
                    "projected_margin": projected_margin,
                    "warnings": warnings,
                }
            )

        preview_rows.sort(
            key=lambda row: (
                0 if row["warnings"] else 1,
                row["projected_profit"] if row["projected_profit"] is not None else -999999.0,
            )
        )
        return render_template(
            "store_settings.html",
            direct_discount_percent=direct_discount_percent,
            min_profit_warning=min_profit_warning,
            min_margin_warning=min_margin_warning,
            listed_count=len(store_items),
            missing_cog_count=missing_cog_count,
            warning_count=warning_count,
            total_direct_profit=total_direct_profit,
            preview_rows=preview_rows,
        )

    @app.post("/settings/store")
    @auth_required
    def store_settings_update():
        fields = [
            ("direct_discount_percent", "Store discount", 0.0, 95.0, "store_direct_discount_percent"),
            ("min_profit_warning", "Minimum profit warning", 0.0, 999999.0, "store_min_profit_warning"),
            ("min_margin_warning", "Minimum margin warning", 0.0, 100.0, "store_min_margin_warning"),
        ]
        parsed = {}
        for form_key, label, min_value, max_value, setting_key in fields:
            raw = request.form.get(form_key, "").strip()
            try:
                value = float(raw)
            except (TypeError, ValueError):
                flash(f"{label} must be a number.", "error")
                return redirect(url_for("store_settings"))
            if value < min_value or value > max_value:
                flash(f"{label} must be between {min_value:g} and {max_value:g}.", "error")
                return redirect(url_for("store_settings"))
            parsed[setting_key] = value

        for key, value in parsed.items():
            _set_app_setting(key, value)
        db.session.commit()
        flash(f"Store settings saved. Direct-buy discount is {parsed['store_direct_discount_percent']:g}%.", "success")
        return redirect(url_for("store_settings"))

    @app.post("/settings/store/item/<int:sku>/direct-price")
    @auth_required
    def store_item_direct_price_update(sku: int):
        item = Item.query.get_or_404(sku)
        raw = request.form.get("direct_price_override", "").strip()
        if raw == "":
            item.direct_price_override = None
            db.session.commit()
            flash(f"Cleared direct price override for SKU {item.sku}.", "success")
            return redirect(url_for("store_settings"))

        direct_price = parse_float(raw)
        if direct_price is None:
            flash("Direct price override must be a valid number.", "error")
            return redirect(url_for("store_settings"))
        if direct_price < 0:
            flash("Direct price override cannot be negative.", "error")
            return redirect(url_for("store_settings"))

        item.direct_price_override = direct_price
        db.session.commit()
        flash(f"Saved direct price override for SKU {item.sku} at ${direct_price:.2f}.", "success")
        return redirect(url_for("store_settings"))

    @app.get("/store")
    def public_store():
        q = request.args.get("q", "").strip()
        category_filter = request.args.get("category", "").strip()
        sort = request.args.get("sort", "newest").strip().lower()
        query = _public_store_query()
        direct_discount_percent = _public_store_discount_percent()

        valid_store_sorts = {"newest", "price_low", "price_high", "discount_amount", "discount_percent"}
        if sort not in valid_store_sorts:
            sort = "newest"

        query = query.order_by(Item.date_listed.desc(), Item.sku.desc())

        all_store_items = _public_store_query().order_by(Item.date_listed.desc(), Item.sku.desc()).all()
        category_order = [
            "Video Games", "Board Games & Toys", "Trading Cards", "Books", "Movies & Music",
            "Clothing", "Electronics", "Collectibles", "Home & Decor", "Sports & Outdoors",
            "Tools & Parts", "Other",
        ]
        category_set = set()
        for item in all_store_items:
            _decorate_public_store_item(item, direct_discount_percent)
            category_set.add(item.store_department)
        categories = [label for label in category_order if label in category_set]

        items = query.all()
        for item in items:
            _decorate_public_store_item(item, direct_discount_percent)
        if q:
            query_text = q.lower()
            items = [
                item for item in items
                if query_text in " ".join([
                    item.item_name or "",
                    item.category or "",
                    item.sub_category or "",
                    item.ebay_category or "",
                    item.ebay_condition or "",
                    item.ebay_item_number or "",
                    item.store_department or "",
                ]).lower()
            ]
        if category_filter:
            items = [item for item in items if item.store_department.lower() == category_filter.lower()]

        if sort == "price_low":
            items.sort(
                key=lambda item: (
                    _public_store_effective_price(item) is None,
                    _public_store_effective_price(item) if _public_store_effective_price(item) is not None else 999999999.0,
                    -(item.sku or 0),
                )
            )
        elif sort == "price_high":
            items.sort(
                key=lambda item: (
                    _public_store_effective_price(item) is not None,
                    _public_store_effective_price(item) if _public_store_effective_price(item) is not None else -1.0,
                    item.sku or 0,
                ),
                reverse=True,
            )
        elif sort == "discount_amount":
            items.sort(
                key=lambda item: (
                    float(item.direct_discount_amount or 0.0),
                    float(item.direct_discount_percent or 0.0),
                    _public_store_effective_price(item) or 0.0,
                    item.sku or 0,
                ),
                reverse=True,
            )
        elif sort == "discount_percent":
            items.sort(
                key=lambda item: (
                    float(item.direct_discount_percent or 0.0),
                    float(item.direct_discount_amount or 0.0),
                    _public_store_effective_price(item) or 0.0,
                    item.sku or 0,
                ),
                reverse=True,
            )
        new_arrivals = all_store_items[:6]
        return render_template(
            "store.html",
            items=items,
            q=q,
            categories=categories,
            category_filter=category_filter,
            sort=sort,
            new_arrivals=new_arrivals,
            direct_discount_percent=direct_discount_percent,
        )

    @app.get("/store/image/<int:image_id>")
    def public_store_image(image_id: int):
        image = ItemImage.query.get_or_404(image_id)
        item = image.item
        if (
            not item or
            item.sold or
            item.canceled or
            not (item.ebay_item_number or "").strip()
        ):
            abort(404)
        return send_from_directory(app.config["UPLOAD_FOLDER"], image.filename)

    def _listed_expr():
        return (
            (Item.ebay_item_number.isnot(None)) &
            (Item.ebay_item_number != "")
        )

    def _needs_listing_info_expr():
        return (
            Item.sold.is_(False) &
            Item.canceled.is_(False) &
            or_(
                ~Item.images.any(),
                Item.cog.is_(None),
                Item.sale_price.is_(None),
                Item.ebay_item_number.is_(None),
                Item.ebay_item_number == "",
            )
        )

    def _inventory_sort_options():
        return {
            "newest": "Newest added",
            "oldest": "Oldest added",
            "az": "A-Z title",
            "za": "Z-A title",
            "date_listed_desc": "Date listed newest",
            "date_listed_asc": "Date listed oldest",
            "date_sold_desc": "Date sold newest",
            "date_sold_asc": "Date sold oldest",
            "price_desc": "Price high to low",
            "price_asc": "Price low to high",
            "profit_desc": "Profit high to low",
            "profit_asc": "Profit low to high",
        }

    def _inventory_query(status_filter="all", platform="", category="", q="", needs_info=False):
        status_filter = (status_filter or "all").strip().lower()
        platform = (platform or "").strip()
        category = (category or "").strip()
        q = (q or "").strip()

        query = Item.query
        listed_expr = _listed_expr()

        if status_filter == "sold":
            query = query.filter(_shipped_sold_expr())
        elif status_filter == "pending_shipping":
            query = query.filter(_pending_shipping_expr())
        elif status_filter == "sold_review":
            query = query.filter(_sold_review_expr())
        elif status_filter == "returned":
            query = query.filter(_returned_expr())
        elif status_filter == "not_listed":
            query = query.filter(Item.sold.is_(False), Item.canceled.is_(False), Item.returned.is_(False)).filter(~listed_expr)
        elif status_filter == "listed":
            query = query.filter(Item.sold.is_(False), Item.canceled.is_(False), Item.returned.is_(False)).filter(listed_expr)
        elif status_filter == "canceled":
            query = query.filter(Item.canceled.is_(True))
        else:
            status_filter = "all"

        if platform:
            query = query.filter(Item.platform == platform)
        if category:
            query = query.filter(Item.category == category)
        if q:
            like = f"%{q}%"
            query = query.filter(
                (Item.item_name.ilike(like)) |
                (Item.notes.ilike(like)) |
                (Item.sub_category.ilike(like)) |
                (Item.category.ilike(like)) |
                (Item.source_location.ilike(like)) |
                (Item.barcode.ilike(like)) |
                (Item.ebay_item_number.ilike(like)) |
                (Item.ebay_order_number.ilike(like)) |
                (Item.ebay_custom_label.ilike(like)) |
                (Item.ebay_item_url.ilike(like)) |
                (Item.ebay_category.ilike(like)) |
                (Item.ebay_condition.ilike(like))
            )

        if needs_info:
            query = query.filter(_needs_listing_info_expr())

        return query, status_filter

    def _apply_inventory_sort(query, sort):
        sort = (sort or "newest").strip().lower()
        sort_options = _inventory_sort_options()
        if sort not in sort_options:
            sort = "newest"

        profit_sort_expr = (
            func.coalesce(Item.buyer_paid_amount, Item.sale_price, 0) -
            func.coalesce(Item.refund_amount, 0) -
            func.coalesce(Item.cog, 0) -
            func.coalesce(Item.shipping, 0) -
            func.coalesce(Item.ad_fee, 0) -
            func.coalesce(Item.ebay_fee, 0)
        )
        if sort == "oldest":
            query = query.order_by(Item.sku.asc())
        elif sort == "az":
            query = query.order_by(func.lower(Item.item_name).asc(), Item.sku.desc())
        elif sort == "za":
            query = query.order_by(func.lower(Item.item_name).desc(), Item.sku.desc())
        elif sort == "date_listed_desc":
            query = query.order_by(Item.date_listed.desc().nullslast(), Item.sku.desc())
        elif sort == "date_listed_asc":
            query = query.order_by(Item.date_listed.asc().nullslast(), Item.sku.desc())
        elif sort == "date_sold_desc":
            query = query.order_by(Item.date_sold.desc().nullslast(), Item.sku.desc())
        elif sort == "date_sold_asc":
            query = query.order_by(Item.date_sold.asc().nullslast(), Item.sku.desc())
        elif sort == "price_desc":
            query = query.order_by(Item.sale_price.desc().nullslast(), Item.sku.desc())
        elif sort == "price_asc":
            query = query.order_by(Item.sale_price.asc().nullslast(), Item.sku.desc())
        elif sort == "profit_desc":
            query = query.order_by(profit_sort_expr.desc(), Item.sku.desc())
        elif sort == "profit_asc":
            query = query.order_by(profit_sort_expr.asc(), Item.sku.desc())
        else:
            query = query.order_by(Item.sku.desc())

        return query, sort

    def _inventory_items_from_url(return_to):
        parsed = urllib.parse.urlparse(return_to or "")
        if parsed.path not in ("", "/"):
            return []

        params = urllib.parse.parse_qs(parsed.query)
        get_param = lambda name, default="": (params.get(name, [default])[-1] or default)
        query, _ = _inventory_query(
            get_param("status", "all"),
            get_param("platform"),
            get_param("category"),
            get_param("q"),
            get_param("needs_info") == "1",
        )
        query, _ = _apply_inventory_sort(query, get_param("sort", "newest"))
        return query.all()

    def _item_neighbors(item, return_to):
        items = _inventory_items_from_url(return_to)
        if not items:
            items = _apply_inventory_sort(Item.query, "newest")[0].all()
        skus = [it.sku for it in items]
        if item.sku not in skus:
            items = _apply_inventory_sort(Item.query, "newest")[0].all()
            skus = [it.sku for it in items]
        if item.sku not in skus:
            return None, None

        index = skus.index(item.sku)
        previous_item = items[index - 1] if index > 0 else None
        next_item = items[index + 1] if index < len(items) - 1 else None
        return previous_item, next_item

    @app.route("/")
    @auth_required
    def index():

        status_filter = request.args.get("status", "all").strip().lower()
        view_mode = request.args.get("view", "cards").strip().lower()
        platform = request.args.get("platform", "").strip()
        category = request.args.get("category", "").strip()
        sort = request.args.get("sort", "newest").strip().lower()
        needs_info = request.args.get("needs_info") == "1"
        q = request.args.get("q", "").strip()

        if view_mode not in {"cards", "table"}:
            view_mode = "cards"

        listed_expr = _listed_expr()
        query, status_filter = _inventory_query(status_filter, platform, category, q, needs_info)
        sort_options = _inventory_sort_options()
        query, sort = _apply_inventory_sort(query, sort)

        items = query.all()

        platforms = get_distinct_values(Item, Item.platform)
        categories = get_distinct_values(Item, Item.category)
        source_locations = get_distinct_values(Item, Item.source_location)
        status_counts = {
            "all": Item.query.count(),
            "not_listed": Item.query.filter(Item.sold.is_(False), Item.canceled.is_(False), Item.returned.is_(False)).filter(~listed_expr).count(),
            "listed": Item.query.filter(Item.sold.is_(False), Item.canceled.is_(False), Item.returned.is_(False)).filter(listed_expr).count(),
            "needs_info": Item.query.filter(_needs_listing_info_expr()).count(),
            "sold_review": Item.query.filter(_sold_review_expr()).count(),
            "pending_shipping": Item.query.filter(_pending_shipping_expr()).count(),
            "sold": Item.query.filter(_shipped_sold_expr()).count(),
            "returned": Item.query.filter(_returned_expr()).count(),
            "canceled": Item.query.filter(Item.canceled.is_(True)).count(),
        }

        return render_template(
            "index.html",
            items=items,
            platforms=platforms,
            categories=categories,
            source_locations=source_locations,
            status_filter=status_filter,
            status_counts=status_counts,
            view_mode=view_mode,
            platform_filter=platform,
            category_filter=category,
            needs_info=needs_info,
            sort=sort,
            sort_options=sort_options,
            q=q,
            current_url=request.full_path.rstrip("?"),
        )

    @app.post("/items/bulk-update")
    @auth_required
    def items_bulk_update():
        raw_skus = request.form.getlist("skus")
        return_to = _safe_return_url(request.form.get("return_to")) or url_for("index")
        source_location_raw = request.form.get("source_location", "")
        cog_raw = request.form.get("cog", "")
        bulk_action = request.form.get("bulk_action", "").strip()
        source_location = source_location_raw.strip()
        cog_text = cog_raw.strip()

        skus = []
        for raw_sku in raw_skus:
            try:
                skus.append(int(raw_sku))
            except (TypeError, ValueError):
                continue

        if not skus:
            flash("Select at least one item to bulk edit.", "warning")
            return redirect(return_to)

        update_location = source_location != ""
        update_cog = cog_text != ""
        update_status = bulk_action in {"sold_review", "shipped"}
        if not update_location and not update_cog and not update_status:
            flash("Enter a location/COG or choose a bulk action before applying.", "warning")
            return redirect(return_to)

        cog_value = None
        if update_cog:
            cog_value = parse_float(cog_text)
            if cog_value is None:
                flash("COG must be a valid number.", "error")
                return redirect(return_to)

        items = Item.query.filter(Item.sku.in_(skus)).all()
        updated = 0
        for item in items:
            changed = False
            if update_location and item.source_location != source_location:
                item.source_location = source_location
                changed = True
            if update_cog and item.cog != cog_value:
                item.cog = cog_value
                changed = True
            if bulk_action == "sold_review":
                if not item.sold or item.sold_confirmed or item.pending_shipping or item.canceled:
                    item.sold = True
                    item.sold_confirmed = False
                    item.pending_shipping = False
                    item.canceled = False
                    if not item.date_sold:
                        item.date_sold = datetime.utcnow().date()
                    changed = True
            elif bulk_action == "shipped":
                if (
                    not item.sold or
                    not item.sold_confirmed or
                    item.pending_shipping or
                    item.canceled or
                    not item.date_shipped
                ):
                    item.sold = True
                    item.sold_confirmed = True
                    item.pending_shipping = False
                    item.canceled = False
                    if not item.date_sold:
                        item.date_sold = datetime.utcnow().date()
                    if not item.date_shipped:
                        item.date_shipped = datetime.utcnow().date()
                    changed = True
            if changed:
                updated += 1

        db.session.commit()
        flash(f"Bulk updated {updated} item{'s' if updated != 1 else ''}.", "success")
        return redirect(return_to)

    @app.route("/categories", methods=["GET", "POST"])
    @auth_required
    def categories_manage():
        def category_expr():
            return func.coalesce(func.nullif(func.trim(Item.category), ""), "Uncategorized")

        if request.method == "POST":
            action = (request.form.get("action") or "rename").strip()
            old_category = (request.form.get("old_category") or "").strip()
            new_category = (request.form.get("new_category") or "").strip()
            sync_ebay_category = request.form.get("sync_ebay_category") == "1"

            if not old_category:
                flash("Pick a category to update.", "warning")
                return redirect(url_for("categories_manage"))

            old_filter = (Item.category.is_(None)) | (func.trim(Item.category) == "") if old_category == "Uncategorized" else (Item.category == old_category)
            items = Item.query.filter(old_filter).all()
            if not items:
                flash("No items matched that category.", "warning")
                return redirect(url_for("categories_manage"))

            if action == "clear":
                replacement = None
            else:
                replacement = new_category or None
                if not replacement:
                    flash("Enter the new category name, or use Clear.", "warning")
                    return redirect(url_for("categories_manage"))

            updated = 0
            for item in items:
                if item.category != replacement:
                    item.category = replacement
                    updated += 1
                if sync_ebay_category and item.ebay_category != replacement:
                    item.ebay_category = replacement
                    updated += 1

            db.session.commit()
            label = replacement or "Uncategorized"
            flash(f"Updated {len(items)} item{'s' if len(items) != 1 else ''}: {old_category} → {label}.", "success")
            return redirect(url_for("categories_manage"))

        cat_col = category_expr()
        returned_expr = _returned_expr()
        financial_item_expr = (
            (Item.sold.is_(True)) &
            (Item.canceled.is_(False)) &
            ((Item.sold_confirmed.is_(True)) | (Item.returned.is_(True)))
        )
        active_unsold_expr = (Item.sold.is_(False)) & (Item.canceled.is_(False)) & (Item.returned.is_(False))
        revenue_expr = case(
            (
                (Item.buyer_paid_amount.isnot(None)) & (Item.sale_price.isnot(None)),
                func.coalesce(Item.buyer_paid_amount, 0.0),
            ),
            (Item.buyer_paid_amount.isnot(None), func.coalesce(Item.buyer_paid_amount, 0.0)),
            (Item.sale_price.isnot(None), func.coalesce(Item.sale_price, 0.0)),
            else_=0.0,
        )
        profit_expr = (
            revenue_expr
            - func.coalesce(Item.refund_amount, 0.0)
            - (
                func.coalesce(Item.cog, 0.0)
                + func.coalesce(Item.shipping, 0.0)
                + func.coalesce(Item.ad_fee, 0.0)
                + func.coalesce(Item.ebay_fee, 0.0)
            )
        )

        rows = (
            db.session.query(
                cat_col.label("category"),
                func.count(Item.sku).label("total_count"),
                func.sum(case((financial_item_expr, 1), else_=0)).label("financial_count"),
                func.sum(case((active_unsold_expr, 1), else_=0)).label("active_count"),
                func.sum(case((returned_expr, 1), else_=0)).label("returned_count"),
                func.coalesce(func.sum(case((financial_item_expr, profit_expr), else_=0.0)), 0.0).label("profit"),
            )
            .group_by(cat_col)
            .order_by(cat_col)
            .all()
        )
        categories = [
            {
                "category": r.category,
                "total_count": int(r.total_count or 0),
                "financial_count": int(r.financial_count or 0),
                "active_count": int(r.active_count or 0),
                "returned_count": int(r.returned_count or 0),
                "profit": float(r.profit or 0.0),
            }
            for r in rows
        ]
        all_category_names = [r["category"] for r in categories]
        return render_template("categories.html", categories=categories, all_category_names=all_category_names)

    @app.route("/reports")
    @auth_required
    def reports():
        range_key = (request.args.get("range") or "this_year").strip().lower()
        start_s = (request.args.get("start") or "").strip()
        end_s = (request.args.get("end") or "").strip()
        trend_mode = (request.args.get("trend") or "month").strip().lower()
        if trend_mode not in {"day", "week", "month"}:
            trend_mode = "month"
        top_n = parse_int(request.args.get("top")) or 10
        top_n = max(5, min(top_n, 25))

        today = datetime.utcnow().date()

        start_date = None
        end_date = None

        if range_key == "30d":
            start_date = today - timedelta(days=30)
            end_date = today
        elif range_key == "90d":
            start_date = today - timedelta(days=90)
            end_date = today
        elif range_key == "this_month":
            start_date = today.replace(day=1)
            end_date = today
        elif range_key == "last_month":
            first_this_month = today.replace(day=1)
            last_month_end = first_this_month - timedelta(days=1)
            start_date = last_month_end.replace(day=1)
            end_date = last_month_end
        elif range_key == "this_year":
            start_date = today.replace(month=1, day=1)
            end_date = today
        elif range_key == "last_year":
            start_date = today.replace(year=today.year - 1, month=1, day=1)
            end_date = today.replace(year=today.year - 1, month=12, day=31)
        elif range_key == "custom":
            start_date = parse_date(start_s)
            end_date = parse_date(end_s)
            if start_date and end_date and start_date > end_date:
                start_date, end_date = end_date, start_date
        else:
            range_key = "all"

        range_labels = {
            "all": "All time",
            "30d": "Last 30 days",
            "90d": "Last 90 days",
            "this_month": "This month",
            "last_month": "Last month",
            "this_year": f"{today.year} year to date",
            "last_year": str(today.year - 1),
            "custom": "Custom range",
        }
        if range_key == "custom" and start_date and end_date:
            selected_range_label = f"{start_date.isoformat()} to {end_date.isoformat()}"
        elif range_key == "custom" and start_date:
            selected_range_label = f"From {start_date.isoformat()}"
        elif range_key == "custom" and end_date:
            selected_range_label = f"Through {end_date.isoformat()}"
        else:
            selected_range_label = range_labels.get(range_key, "All time")

        sold_date_filters = []
        if start_date:
            sold_date_filters.append(Item.date_sold.isnot(None))
            sold_date_filters.append(Item.date_sold >= start_date)
        if end_date:
            sold_date_filters.append(Item.date_sold.isnot(None))
            sold_date_filters.append(Item.date_sold <= end_date)

        def nz(col):
            return func.coalesce(col, 0.0)

        buyer_shipping_paid_expr = case(
            (
                (Item.buyer_paid_amount.isnot(None)) & (Item.sale_price.isnot(None)),
                func.coalesce(Item.buyer_paid_amount, 0.0) - func.coalesce(Item.sale_price, 0.0),
            ),
            else_=0.0,
        )
        report_revenue_expr = case(
            (
                (Item.buyer_paid_amount.isnot(None)) & (Item.sale_price.isnot(None)),
                func.coalesce(Item.sale_price, 0.0) + buyer_shipping_paid_expr,
            ),
            (
                Item.buyer_paid_amount.isnot(None),
                func.coalesce(Item.buyer_paid_amount, 0.0),
            ),
            (
                Item.sale_price.isnot(None),
                func.coalesce(Item.sale_price, 0.0),
            ),
            else_=0.0,
        )
        profit_expr = (
            report_revenue_expr
            - nz(Item.refund_amount)
            - (nz(Item.cog) + nz(Item.shipping) + nz(Item.ad_fee) + nz(Item.ebay_fee))
        )

        days_to_sell_expr = case(
            (
                (Item.date_listed.isnot(None)) & (Item.date_sold.isnot(None)),
                func.julianday(Item.date_sold) - func.julianday(Item.date_listed),
            ),
            else_=None,
        )

        category_col = func.coalesce(Item.category, "Uncategorized")
        source_col = func.coalesce(Item.source_location, "Unknown")
        confirmed_sold_expr = (
            (Item.sold.is_(True)) &
            (Item.sold_confirmed.is_(True)) &
            (Item.canceled.is_(False)) &
            (Item.returned.is_(False))
        )
        returned_expr = _returned_expr()
        financial_item_expr = (
            (Item.sold.is_(True)) &
            (Item.canceled.is_(False)) &
            ((Item.sold_confirmed.is_(True)) | (Item.returned.is_(True)))
        )
        active_unsold_expr = (Item.sold.is_(False)) & (Item.canceled.is_(False)) & (Item.returned.is_(False))
        listed_expr = (Item.ebay_item_number.isnot(None)) & (Item.ebay_item_number != "")

        total_items = Item.query.count()
        listed_items = Item.query.filter(active_unsold_expr, listed_expr).count()
        not_listed_items = Item.query.filter(active_unsold_expr).filter(~listed_expr).count()
        sold_review_items = Item.query.filter(_sold_review_expr()).count()
        pending_shipping_items = Item.query.filter(_pending_shipping_expr()).count()
        returned_items = Item.query.filter(returned_expr).count()
        canceled_items = Item.query.filter(Item.canceled.is_(True)).count()
        refund_total_q = db.session.query(func.coalesce(func.sum(Item.refund_amount), 0.0)).filter(returned_expr)
        if start_date:
            refund_total_q = refund_total_q.filter(Item.date_returned.isnot(None), Item.date_returned >= start_date)
        if end_date:
            refund_total_q = refund_total_q.filter(Item.date_returned.isnot(None), Item.date_returned <= end_date)
        refund_total = float(refund_total_q.scalar() or 0.0)

        unsold_inventory_cost = float(
            db.session.query(func.coalesce(func.sum(Item.cog), 0.0))
            .filter(active_unsold_expr)
            .scalar() or 0.0
        )
        unsold_listed_value = float(
            db.session.query(func.coalesce(func.sum(Item.sale_price), 0.0))
            .filter(active_unsold_expr, listed_expr)
            .scalar() or 0.0
        )
        listed_inventory_cost = float(
            db.session.query(func.coalesce(func.sum(Item.cog), 0.0))
            .filter(active_unsold_expr, listed_expr)
            .scalar() or 0.0
        )
        listed_potential_profit = unsold_listed_value - listed_inventory_cost
        listed_inventory_markup_pct = (listed_potential_profit / listed_inventory_cost * 100.0) if listed_inventory_cost else 0.0
        listed_inventory_margin_pct = (listed_potential_profit / unsold_listed_value * 100.0) if unsold_listed_value else 0.0
        listed_inventory_scale = max(unsold_listed_value, listed_inventory_cost, abs(listed_potential_profit), 1.0)

        missing_sold_review = {
            "buyer_paid": Item.query.filter(_sold_review_expr(), Item.buyer_paid_amount.is_(None)).count(),
            "cog": Item.query.filter(_sold_review_expr(), Item.cog.is_(None)).count(),
            "shipping": Item.query.filter(_sold_review_expr(), Item.shipping.is_(None)).count(),
            "ad_fee": Item.query.filter(_sold_review_expr(), Item.ad_fee.is_(None)).count(),
            "ebay_fee": Item.query.filter(_sold_review_expr(), Item.ebay_fee.is_(None)).count(),
        }

        sold_items_q = Item.query.filter(confirmed_sold_expr)
        if sold_date_filters:
            sold_items_q = sold_items_q.filter(*sold_date_filters)
        sold_items = sold_items_q.count()

        financial_items_q = Item.query.filter(financial_item_expr)
        if sold_date_filters:
            financial_items_q = financial_items_q.filter(*sold_date_filters)
        financial_items = financial_items_q.count()

        sold_rate_pct = (sold_items / total_items * 100.0) if total_items else 0.0

        total_profit_q = (
            db.session.query(func.coalesce(func.sum(profit_expr), 0.0))
            .filter(financial_item_expr)
        )
        if sold_date_filters:
            total_profit_q = total_profit_q.filter(*sold_date_filters)
        total_profit = float(total_profit_q.scalar() or 0.0)

        ytd_start = today.replace(month=1, day=1)
        ytd_profit = float(
            db.session.query(func.coalesce(func.sum(profit_expr), 0.0))
            .filter(financial_item_expr, Item.date_sold.isnot(None), Item.date_sold >= ytd_start, Item.date_sold <= today)
            .scalar() or 0.0
        )
        ytd_sold_items = int(
            Item.query
            .filter(confirmed_sold_expr, Item.date_sold.isnot(None), Item.date_sold >= ytd_start, Item.date_sold <= today)
            .count()
        )
        avg_profit_per_sold = (total_profit / financial_items) if financial_items else 0.0
        shipping_shortfall_expr = case(
            (
                (Item.buyer_paid_amount.isnot(None)) &
                (Item.sale_price.isnot(None)) &
                (func.coalesce(Item.shipping, 0.0) > buyer_shipping_paid_expr),
                func.coalesce(Item.shipping, 0.0) - buyer_shipping_paid_expr,
            ),
            else_=0.0,
        )
        shipping_shortfall_count_expr = case(
            (
                (Item.buyer_paid_amount.isnot(None)) &
                (Item.sale_price.isnot(None)) &
                (func.coalesce(Item.shipping, 0.0) > buyer_shipping_paid_expr),
                1,
            ),
            else_=0,
        )
        financial_totals_q = (
            db.session.query(
                func.coalesce(func.sum(report_revenue_expr), 0.0).label("gross_sales"),
                func.coalesce(func.sum(buyer_shipping_paid_expr), 0.0).label("buyer_shipping_paid"),
                func.coalesce(func.sum(Item.cog), 0.0).label("cog"),
                func.coalesce(func.sum(Item.shipping), 0.0).label("shipping"),
                func.coalesce(func.sum(shipping_shortfall_expr), 0.0).label("shipping_shortfall"),
                func.coalesce(func.sum(shipping_shortfall_count_expr), 0).label("shipping_shortfall_count"),
                func.coalesce(func.sum(Item.ebay_fee), 0.0).label("ebay_fee"),
                func.coalesce(func.sum(Item.ad_fee), 0.0).label("ad_fee"),
            )
            .filter(financial_item_expr)
        )
        if sold_date_filters:
            financial_totals_q = financial_totals_q.filter(*sold_date_filters)
        financial_totals_row = financial_totals_q.one()
        gross_sales_total = float(financial_totals_row.gross_sales or 0.0)
        buyer_shipping_paid_total = float(financial_totals_row.buyer_shipping_paid or 0.0)
        cog_total = float(financial_totals_row.cog or 0.0)
        shipping_total = float(financial_totals_row.shipping or 0.0)
        shipping_shortfall_total = float(financial_totals_row.shipping_shortfall or 0.0)
        shipping_shortfall_count = int(financial_totals_row.shipping_shortfall_count or 0)
        shipping_net_total = buyer_shipping_paid_total - shipping_total
        shipping_recovery_pct = (buyer_shipping_paid_total / shipping_total * 100.0) if shipping_total else 0.0
        ebay_fee_total = float(financial_totals_row.ebay_fee or 0.0)
        ad_fee_total = float(financial_totals_row.ad_fee or 0.0)
        operating_cost_total = cog_total + shipping_total + ebay_fee_total + ad_fee_total
        money_out_total = operating_cost_total + refund_total
        net_sales_total = gross_sales_total - refund_total
        profit_margin_pct = (total_profit / gross_sales_total * 100.0) if gross_sales_total else 0.0

        avg_days_to_sell_q = (
            db.session.query(func.avg(days_to_sell_expr))
            .filter(confirmed_sold_expr)
        )
        if sold_date_filters:
            avg_days_to_sell_q = avg_days_to_sell_q.filter(*sold_date_filters)
        avg_days_to_sell = avg_days_to_sell_q.scalar()
        avg_days_to_sell = float(avg_days_to_sell) if avg_days_to_sell is not None else 0.0

        # By Category (existing)
        sold_count_all = func.sum(case((confirmed_sold_expr, 1), else_=0))
        unsold_count = func.sum(case(((Item.sold.is_(False)) & (Item.canceled.is_(False)), 1), else_=0))
        total_count = func.count(Item.sku)

        avg_days_listed_unsold = func.avg(
            case(
                (
                    (Item.sold.is_(False)) & (Item.canceled.is_(False)) & (Item.date_listed.isnot(None)),
                    func.julianday(func.current_date()) - func.julianday(Item.date_listed),
                ),
                else_=None,
            )
        )

        rows_counts = (
            db.session.query(
                category_col.label("category"),
                sold_count_all.label("sold_count_all"),
                unsold_count.label("unsold_count"),
                total_count.label("total_count"),
                avg_days_listed_unsold.label("avg_days_listed_unsold"),
            )
            .group_by(category_col)
            .all()
        )

        counts_map = {}
        for r in rows_counts:
            counts_map[r.category] = {
                "category": r.category,
                "unsold_count": int(r.unsold_count or 0),
                "total_count": int(r.total_count or 0),
                "avg_days_listed_unsold": float(r.avg_days_listed_unsold) if r.avg_days_listed_unsold is not None else None,
            }

        sold_metrics_q = (
            db.session.query(
                category_col.label("category"),
                func.count(Item.sku).label("sold_count"),
                func.coalesce(func.sum(profit_expr), 0.0).label("total_profit"),
                func.avg(profit_expr).label("avg_profit"),
            )
            .filter(financial_item_expr)
        )
        if sold_date_filters:
            sold_metrics_q = sold_metrics_q.filter(*sold_date_filters)
        sold_rows = sold_metrics_q.group_by(category_col).all()

        sold_map = {}
        for r in sold_rows:
            sold_map[r.category] = {
                "sold_count": int(r.sold_count or 0),
                "total_profit": float(r.total_profit or 0.0),
                "avg_profit": float(r.avg_profit) if r.avg_profit is not None else 0.0,
            }

        by_category = []
        all_cats = sorted(set(list(counts_map.keys()) + list(sold_map.keys())))
        for cat in all_cats:
            c = counts_map.get(cat, {"unsold_count": 0, "total_count": 0, "avg_days_listed_unsold": None})
            s = sold_map.get(cat, {"sold_count": 0, "total_profit": 0.0, "avg_profit": 0.0})

            total_count_val = int(c.get("total_count") or 0)
            sold_count_val = int(s.get("sold_count") or 0)
            unsold_count_val = int(c.get("unsold_count") or 0)
            sold_rate_pct_cat = (sold_count_val * 100.0 / total_count_val) if total_count_val else 0.0

            by_category.append(
                {
                    "category": cat,
                    "sold_count": sold_count_val,
                    "unsold_count": unsold_count_val,
                    "sold_rate_pct": float(sold_rate_pct_cat),
                    "total_profit": float(s.get("total_profit") or 0.0),
                    "avg_profit": float(s.get("avg_profit") or 0.0),
                    "avg_days_listed_unsold": c.get("avg_days_listed_unsold"),
                }
            )
        by_category.sort(key=lambda x: (x["sold_count"], x["total_profit"]), reverse=True)

        # By Source Location (NEW)
        sold_count_src = func.sum(case((confirmed_sold_expr, 1), else_=0))
        unsold_count_src = func.sum(case(((Item.sold.is_(False)) & (Item.canceled.is_(False)), 1), else_=0))
        total_count_src = func.count(Item.sku)

        sold_profit_src = func.coalesce(func.sum(case((confirmed_sold_expr, profit_expr), else_=0.0)), 0.0)
        avg_profit_src = func.avg(case((confirmed_sold_expr, profit_expr), else_=None))
        avg_days_to_sell_src = func.avg(case((confirmed_sold_expr, days_to_sell_expr), else_=None))

        avg_days_listed_unsold_src = func.avg(
            case(
                (
                    (Item.sold.is_(False)) & (Item.canceled.is_(False)) & (Item.date_listed.isnot(None)),
                    func.julianday(func.current_date()) - func.julianday(Item.date_listed),
                ),
                else_=None,
            )
        )

        avg_cog_unsold_src = func.avg(
            case(
                ((Item.sold.is_(False)) & (Item.canceled.is_(False)) & (Item.cog.isnot(None)), Item.cog),
                else_=None,
            )
        )

        src_base_q = db.session.query(
            source_col.label("source"),
            sold_count_src.label("sold_count_all"),
            unsold_count_src.label("unsold_count"),
            total_count_src.label("total_count"),
            avg_days_listed_unsold_src.label("avg_days_listed_unsold"),
            avg_cog_unsold_src.label("avg_cog_unsold"),
        ).group_by(source_col)

        src_rows_counts = src_base_q.all()
        src_counts_map = {}
        for r in src_rows_counts:
            src_counts_map[r.source] = {
                "source": r.source,
                "unsold_count": int(r.unsold_count or 0),
                "total_count": int(r.total_count or 0),
                "avg_days_listed_unsold": float(r.avg_days_listed_unsold) if r.avg_days_listed_unsold is not None else None,
                "avg_cog_unsold": float(r.avg_cog_unsold) if r.avg_cog_unsold is not None else None,
            }

        src_sold_q = db.session.query(
            source_col.label("source"),
            func.count(Item.sku).label("sold_count"),
            func.coalesce(func.sum(profit_expr), 0.0).label("total_profit"),
            func.avg(profit_expr).label("avg_profit"),
            func.avg(days_to_sell_expr).label("avg_days_to_sell"),
        ).filter(financial_item_expr)

        if sold_date_filters:
            src_sold_q = src_sold_q.filter(*sold_date_filters)

        src_sold_rows = src_sold_q.group_by(source_col).all()
        src_sold_map = {}
        for r in src_sold_rows:
            src_sold_map[r.source] = {
                "sold_count": int(r.sold_count or 0),
                "total_profit": float(r.total_profit or 0.0),
                "avg_profit": float(r.avg_profit) if r.avg_profit is not None else 0.0,
                "avg_days_to_sell": float(r.avg_days_to_sell) if r.avg_days_to_sell is not None else None,
            }

        by_source = []
        all_sources = sorted(set(list(src_counts_map.keys()) + list(src_sold_map.keys())))
        for src in all_sources:
            c = src_counts_map.get(src, {"unsold_count": 0, "total_count": 0, "avg_days_listed_unsold": None, "avg_cog_unsold": None})
            s = src_sold_map.get(src, {"sold_count": 0, "total_profit": 0.0, "avg_profit": 0.0, "avg_days_to_sell": None})

            total_count_val = int(c.get("total_count") or 0)
            sold_count_val = int(s.get("sold_count") or 0)
            unsold_count_val = int(c.get("unsold_count") or 0)
            sold_rate_pct_src = (sold_count_val * 100.0 / total_count_val) if total_count_val else 0.0

            by_source.append(
                {
                    "source": src,
                    "sold_count": sold_count_val,
                    "unsold_count": unsold_count_val,
                    "sold_rate_pct": float(sold_rate_pct_src),
                    "total_profit": float(s.get("total_profit") or 0.0),
                    "avg_profit": float(s.get("avg_profit") or 0.0),
                    "avg_days_to_sell": s.get("avg_days_to_sell"),
                    "avg_days_listed_unsold": c.get("avg_days_listed_unsold"),
                    "avg_cog_unsold": c.get("avg_cog_unsold"),
                }
            )

        by_source.sort(key=lambda x: (x["sold_count"], x["total_profit"]), reverse=True)

        best_source = next((r for r in sorted(by_source, key=lambda x: (x["total_profit"], x["sold_count"]), reverse=True) if r["sold_count"]), None)
        best_category = next((r for r in sorted(by_category, key=lambda x: (x["total_profit"], x["sold_count"]), reverse=True) if r["sold_count"]), None)

        pipeline_items = [
            {"label": "Listed", "count": listed_items, "class": "listed", "href": url_for("index", status="listed")},
            {"label": "Not listed", "count": not_listed_items, "class": "not-listed", "href": url_for("index", status="not_listed")},
            {"label": "Sold review", "count": sold_review_items, "class": "review", "href": url_for("index", status="sold_review")},
            {"label": "Pending shipping", "count": pending_shipping_items, "class": "pending", "href": url_for("index", status="pending_shipping")},
            {"label": "Shipped sold", "count": Item.query.filter(_shipped_sold_expr()).count(), "class": "sold", "href": url_for("index", status="sold")},
            {"label": "Refunded", "count": returned_items, "class": "returned", "href": url_for("index", status="returned")},
            {"label": "Canceled", "count": canceled_items, "class": "canceled", "href": url_for("index", status="canceled")},
        ]
        pipeline_total = sum(p["count"] for p in pipeline_items) or 1
        for p in pipeline_items:
            p["pct"] = (p["count"] * 100.0 / pipeline_total) if pipeline_total else 0.0

        source_chart = sorted(by_source, key=lambda x: (x["total_profit"], x["sold_count"]), reverse=True)[:8]
        max_source_profit = max([abs(r["total_profit"]) for r in source_chart] + [1.0])
        for r in source_chart:
            r["profit_width"] = max(4.0, abs(r["total_profit"]) * 100.0 / max_source_profit)

        category_chart = sorted(by_category, key=lambda x: (x["total_profit"], x["sold_count"]), reverse=True)[:8]
        max_category_profit = max([abs(r["total_profit"]) for r in category_chart] + [1.0])
        for r in category_chart:
            r["profit_width"] = max(4.0, abs(r["total_profit"]) * 100.0 / max_category_profit)

        top_category_names = [r["category"] for r in category_chart[:5] if r.get("category")]
        category_line_chart = {
            "periods": [],
            "series": [],
            "baseline_y": 50,
        }
        category_season_rows = []
        if top_category_names:
            category_month_expr = func.strftime("%Y-%m", Item.date_sold)
            category_line_q = (
                db.session.query(
                    category_col.label("category"),
                    category_month_expr.label("period"),
                    func.coalesce(func.sum(profit_expr), 0.0).label("profit"),
                )
                .filter(financial_item_expr, Item.date_sold.isnot(None), category_col.in_(top_category_names))
            )
            if sold_date_filters:
                category_line_q = category_line_q.filter(*sold_date_filters)
            category_line_rows = category_line_q.group_by(category_col, category_month_expr).order_by(category_month_expr.asc()).all()
            category_periods = sorted({r.period for r in category_line_rows if r.period})[-12:]
            category_period_labels = []
            for period in category_periods:
                try:
                    category_period_labels.append(datetime.strptime(period, "%Y-%m").strftime("%b"))
                except ValueError:
                    category_period_labels.append(period)

            profit_by_category_period = {}
            for r in category_line_rows:
                if r.period in category_periods:
                    profit_by_category_period[(r.category, r.period)] = float(r.profit or 0.0)
            max_category_line_profit = max([abs(v) for v in profit_by_category_period.values()] + [1.0])
            chart_colors = ["#60a5fa", "#22c55e", "#fbbf24", "#f87171", "#c084fc"]
            series = []
            for idx, category_name in enumerate(top_category_names):
                values = [profit_by_category_period.get((category_name, period), 0.0) for period in category_periods]
                points = []
                for point_idx, value in enumerate(values):
                    x = 50.0 if len(category_periods) == 1 else point_idx * 100.0 / max(len(category_periods) - 1, 1)
                    y = 50.0 - (value * 45.0 / max_category_line_profit)
                    y = max(5.0, min(95.0, y))
                    points.append(f"{x:.2f},{y:.2f}")
                if points:
                    series.append({
                        "category": category_name,
                        "color": chart_colors[idx % len(chart_colors)],
                        "points": " ".join(points),
                        "latest_profit": values[-1] if values else 0.0,
                    })
            category_line_chart = {
                "periods": category_periods,
                "labels": category_period_labels,
                "series": series,
                "baseline_y": 50,
            }

            season_q = (
                db.session.query(
                    category_col.label("category"),
                    func.strftime("%m", Item.date_sold).label("month"),
                    func.coalesce(func.sum(profit_expr), 0.0).label("profit"),
                    func.count(Item.sku).label("count"),
                )
                .filter(financial_item_expr, Item.date_sold.isnot(None), category_col.in_(top_category_names))
            )
            if sold_date_filters:
                season_q = season_q.filter(*sold_date_filters)
            season_rows_raw = season_q.group_by(category_col, func.strftime("%m", Item.date_sold)).all()
            season_order = ["Winter", "Spring", "Summer", "Fall"]
            season_months = {
                "12": "Winter", "01": "Winter", "02": "Winter",
                "03": "Spring", "04": "Spring", "05": "Spring",
                "06": "Summer", "07": "Summer", "08": "Summer",
                "09": "Fall", "10": "Fall", "11": "Fall",
            }
            season_map = {}
            for r in season_rows_raw:
                season = season_months.get(r.month or "", "Other")
                key = (r.category, season)
                current = season_map.setdefault(key, {"profit": 0.0, "count": 0})
                current["profit"] += float(r.profit or 0.0)
                current["count"] += int(r.count or 0)

            max_season_profit = max([abs(v["profit"]) for v in season_map.values()] + [1.0])
            for category_name in top_category_names:
                cells = []
                total_profit_for_category = 0.0
                for season in season_order:
                    data = season_map.get((category_name, season), {"profit": 0.0, "count": 0})
                    total_profit_for_category += data["profit"]
                    cells.append({
                        "season": season,
                        "profit": data["profit"],
                        "count": data["count"],
                        "width": max(4.0, abs(data["profit"]) * 100.0 / max_season_profit) if data["profit"] else 0.0,
                    })
                category_season_rows.append({
                    "category": category_name,
                    "total_profit": total_profit_for_category,
                    "cells": cells,
                })
            category_season_rows.sort(key=lambda x: x["total_profit"], reverse=True)

        if trend_mode == "day":
            trend_period_expr = func.strftime("%Y-%m-%d", Item.date_sold)
            trend_limit = 30
            trend_title = "Daily profit"
            trend_note = "Last 30 sold days in range"
        elif trend_mode == "week":
            trend_period_expr = func.strftime("%Y-W%W", Item.date_sold)
            trend_limit = 26
            trend_title = "Weekly profit"
            trend_note = "Last 26 sold weeks in range"
        else:
            trend_period_expr = func.strftime("%Y-%m", Item.date_sold)
            trend_limit = 12
            trend_title = "Monthly profit"
            trend_note = "Last 12 sold months in range"

        trend_q = (
            db.session.query(
                trend_period_expr.label("period"),
                func.coalesce(func.sum(profit_expr), 0.0).label("profit"),
                func.coalesce(func.sum(report_revenue_expr), 0.0).label("gross_sales"),
                func.coalesce(func.sum(Item.shipping), 0.0).label("shipping"),
                func.coalesce(func.sum(Item.ebay_fee), 0.0).label("ebay_fee"),
                func.coalesce(func.sum(Item.ad_fee), 0.0).label("ad_fee"),
                func.count(Item.sku).label("sold_count"),
            )
            .filter(financial_item_expr, Item.date_sold.isnot(None))
        )
        if sold_date_filters:
            trend_q = trend_q.filter(*sold_date_filters)
        trend_rows = trend_q.group_by("period").order_by(text("period DESC")).limit(trend_limit).all()
        profit_trend = [
            {
                "period": r.period,
                "label": r.period,
                "profit": float(r.profit or 0.0),
                "gross_sales": float(r.gross_sales or 0.0),
                "shipping": float(r.shipping or 0.0),
                "ebay_fee": float(r.ebay_fee or 0.0),
                "ad_fee": float(r.ad_fee or 0.0),
                "sold_count": int(r.sold_count or 0),
            }
            for r in reversed(trend_rows)
        ]
        for idx, row in enumerate(profit_trend):
            if trend_mode == "day":
                parsed_period = parse_date(row["period"])
                row["label"] = f"{parsed_period.month}/{parsed_period.day}" if parsed_period else row["period"]
                row["show_label"] = idx % 3 == 0 or idx == len(profit_trend) - 1
            elif trend_mode == "week":
                row["label"] = row["period"].replace("-W", " W")
                row["show_label"] = idx % 2 == 0 or idx == len(profit_trend) - 1
            else:
                try:
                    parsed_period = datetime.strptime(row["period"], "%Y-%m")
                    row["label"] = parsed_period.strftime("%b %Y")
                except ValueError:
                    row["label"] = row["period"]
                row["show_label"] = True

        max_trend_profit = max([math.sqrt(abs(r["profit"])) for r in profit_trend if r["profit"]] + [1.0])
        for r in profit_trend:
            scaled = math.sqrt(abs(r["profit"])) if r["profit"] else 0.0
            r["height"] = max(8.0, scaled * 100.0 / max_trend_profit) if scaled else 3.0

        cost_breakdown = [
            {"label": "COG", "amount": cog_total, "class": "cog"},
            {"label": "Shipping", "amount": shipping_total, "class": "shipping"},
            {"label": "eBay fees", "amount": ebay_fee_total, "class": "fees"},
            {"label": "Ad fees", "amount": ad_fee_total, "class": "ads"},
            {"label": "Refunds", "amount": refund_total, "class": "refunds"},
        ]
        cost_scale_total = money_out_total or max([row["amount"] for row in cost_breakdown] + [1.0])
        for row in cost_breakdown:
            row["share_pct"] = (row["amount"] * 100.0 / cost_scale_total) if cost_scale_total else 0.0
            row["pct"] = max(4.0, row["share_pct"]) if row["amount"] else 0.0

        # Top profit items (sold in range)
        top_q = (
            db.session.query(
                Item.sku,
                Item.item_name,
                category_col.label("category"),
                profit_expr.label("profit"),
                days_to_sell_expr.label("days_to_sell"),
                Item.date_sold.label("date_sold"),
            )
            .filter(financial_item_expr)
        )
        if sold_date_filters:
            top_q = top_q.filter(*sold_date_filters)

        top_rows = top_q.order_by(profit_expr.desc()).limit(top_n).all()

        top_profit = []
        for r in top_rows:
            top_profit.append(
                {
                    "sku": r.sku,
                    "item_name": r.item_name,
                    "category": r.category,
                    "profit": float(r.profit or 0.0),
                    "days_to_sell": float(r.days_to_sell) if r.days_to_sell is not None else None,
                    "date_sold": r.date_sold.isoformat() if r.date_sold else None,
                }
            )
                # --- attach thumbnail urls for top_profit ---
        try:
            top_skus = [tp["sku"] for tp in top_profit if tp.get("sku") is not None]
            if top_skus:
                # Pull first image per SKU (simple approach)
                imgs = (
                    ItemImage.query
                    .filter(ItemImage.item_sku.in_(top_skus))
                    .order_by(ItemImage.item_sku.asc(), ItemImage.id.asc())
                    .all()
                )

                first_img = {}
                for im in imgs:
                    if im.item_sku not in first_img and im.filename:
                        first_img[im.item_sku] = im.filename

                for tp in top_profit:
                    fn = first_img.get(tp["sku"])
                    tp["thumb_url"] = url_for("uploaded_file", filename=fn) if fn else ""
        except Exception as e:
            # don't let reports page crash if something goes sideways
            for tp in top_profit:
                tp["thumb_url"] = ""

        sold_review_rows = (
            Item.query
            .filter(_sold_review_expr())
            .order_by(Item.date_sold.desc(), Item.sku.desc())
            .limit(12)
            .all()
        )
        sold_review_attention = []
        for item in sold_review_rows:
            missing = []
            if item.buyer_paid_amount is None:
                missing.append("Buyer paid")
            if item.cog is None:
                missing.append("COG")
            if item.shipping is None:
                missing.append("Shipping")
            if item.ad_fee is None:
                missing.append("Ad fee")
            if item.ebay_fee is None:
                missing.append("eBay fee")
            sold_review_attention.append({
                "sku": item.sku,
                "item_name": item.item_name,
                "date_sold": item.date_sold.isoformat() if item.date_sold else "",
                "missing": missing,
                "profit": float(item.profit or 0.0),
            })

        aged_unsold_rows = (
            Item.query
            .filter(active_unsold_expr, Item.date_listed.isnot(None))
            .order_by(Item.date_listed.asc(), Item.sku.desc())
            .limit(12)
            .all()
        )
        aged_unsold = []
        for item in aged_unsold_rows:
            days_listed = (today - item.date_listed).days if item.date_listed else None
            aged_unsold.append({
                "sku": item.sku,
                "item_name": item.item_name,
                "date_listed": item.date_listed.isoformat() if item.date_listed else "",
                "days_listed": days_listed,
                "sale_price": float(item.sale_price or 0.0),
                "listed": bool(item.ebay_item_number),
            })

        negative_q = (
            db.session.query(
                Item.sku,
                Item.item_name,
                report_revenue_expr.label("revenue"),
                Item.cog.label("cog"),
                Item.shipping.label("shipping"),
                Item.ebay_fee.label("ebay_fee"),
                Item.ad_fee.label("ad_fee"),
                Item.refund_amount.label("refund_amount"),
                profit_expr.label("profit"),
                Item.date_sold.label("date_sold"),
            )
            .filter(financial_item_expr)
            .filter(profit_expr < 0)
        )
        if sold_date_filters:
            negative_q = negative_q.filter(*sold_date_filters)
        negative_rows = negative_q.order_by(profit_expr.asc()).limit(12).all()
        negative_profit = []
        for r in negative_rows:
            revenue = float(r.revenue or 0.0)
            cog = float(r.cog or 0.0)
            shipping = float(r.shipping or 0.0)
            ebay_fee = float(r.ebay_fee or 0.0)
            ad_fee = float(r.ad_fee or 0.0)
            refund_amount = float(r.refund_amount or 0.0)
            biggest_cost_label = "COG"
            biggest_cost_amount = cog
            for label, amount in [("Shipping", shipping), ("eBay fees", ebay_fee), ("Ad fees", ad_fee), ("Refund", refund_amount)]:
                if amount > biggest_cost_amount:
                    biggest_cost_label = label
                    biggest_cost_amount = amount
            negative_profit.append(
                {
                    "sku": r.sku,
                    "item_name": r.item_name,
                    "revenue": revenue,
                    "cog": cog,
                    "shipping": shipping,
                    "ebay_fee": ebay_fee,
                    "ad_fee": ad_fee,
                    "refund_amount": refund_amount,
                    "profit": float(r.profit or 0.0),
                    "loss_amount": abs(float(r.profit or 0.0)),
                    "biggest_cost_label": biggest_cost_label,
                    "biggest_cost_amount": biggest_cost_amount,
                    "date_sold": r.date_sold.isoformat() if r.date_sold else "",
                }
            )

        negative_profit_count_q = db.session.query(func.count(Item.sku)).filter(financial_item_expr, profit_expr < 0)
        negative_profit_total_q = db.session.query(func.coalesce(func.sum(profit_expr), 0.0)).filter(financial_item_expr, profit_expr < 0)
        if sold_date_filters:
            negative_profit_count_q = negative_profit_count_q.filter(*sold_date_filters)
            negative_profit_total_q = negative_profit_total_q.filter(*sold_date_filters)
        negative_profit_count = int(negative_profit_count_q.scalar() or 0)
        negative_profit_total = float(negative_profit_total_q.scalar() or 0.0)

        returned_rows_q = (
            Item.query
            .filter(returned_expr)
            .order_by(Item.date_returned.desc().nullslast(), Item.date_sold.desc().nullslast(), Item.sku.desc())
        )
        if start_date:
            returned_rows_q = returned_rows_q.filter(Item.date_returned.isnot(None), Item.date_returned >= start_date)
        if end_date:
            returned_rows_q = returned_rows_q.filter(Item.date_returned.isnot(None), Item.date_returned <= end_date)
        returned_attention = []
        for item in returned_rows_q.limit(12).all():
            returned_attention.append({
                "sku": item.sku,
                "item_name": item.item_name,
                "date_returned": item.date_returned.isoformat() if item.date_returned else "",
                "refund_amount": float(item.refund_amount or 0.0),
                "profit": float(item.profit or 0.0),
                "return_reference_id": item.return_reference_id or "",
            })

        kpis = {
            "total_items": total_items,
            "listed_items": listed_items,
            "not_listed_items": not_listed_items,
            "sold_review_items": sold_review_items,
            "pending_shipping_items": pending_shipping_items,
            "returned_items": returned_items,
            "refund_total": refund_total,
            "canceled_items": canceled_items,
            "sold_items": sold_items,
            "financial_items": financial_items,
            "sold_rate_pct": sold_rate_pct,
            "total_profit": float(total_profit),
            "selected_range_label": selected_range_label,
            "ytd_profit": ytd_profit,
            "ytd_sold_items": ytd_sold_items,
            "ytd_year": today.year,
            "avg_profit_per_sold": float(avg_profit_per_sold),
            "gross_sales_total": gross_sales_total,
            "net_sales_total": net_sales_total,
            "cog_total": cog_total,
            "shipping_total": shipping_total,
            "buyer_shipping_paid_total": buyer_shipping_paid_total,
            "shipping_net_total": shipping_net_total,
            "shipping_shortfall_total": shipping_shortfall_total,
            "shipping_shortfall_count": shipping_shortfall_count,
            "shipping_recovery_pct": shipping_recovery_pct,
            "negative_profit_count": negative_profit_count,
            "negative_profit_total": negative_profit_total,
            "ebay_fee_total": ebay_fee_total,
            "ad_fee_total": ad_fee_total,
            "operating_cost_total": operating_cost_total,
            "money_out_total": money_out_total,
            "profit_margin_pct": profit_margin_pct,
            "avg_days_to_sell": float(avg_days_to_sell),
            "unsold_inventory_cost": unsold_inventory_cost,
            "unsold_listed_value": unsold_listed_value,
            "listed_inventory_cost": listed_inventory_cost,
            "listed_potential_profit": listed_potential_profit,
            "listed_inventory_markup_pct": listed_inventory_markup_pct,
            "listed_inventory_margin_pct": listed_inventory_margin_pct,
            "listed_inventory_scale": listed_inventory_scale,
            "missing_sold_review": missing_sold_review,
            "best_source": best_source,
            "best_category": best_category,
        }

        report_viz = {
            "pipeline_items": pipeline_items,
            "pipeline_total": pipeline_total,
            "source_chart": source_chart,
            "category_chart": category_chart,
            "category_line_chart": category_line_chart,
            "category_season_rows": category_season_rows,
            "profit_trend": profit_trend,
            "cost_breakdown": cost_breakdown,
            "trend_title": trend_title,
            "trend_note": trend_note,
        }

        return render_template(
            "reports.html",
            kpis=kpis,
            by_category=by_category,
            by_source=by_source,
            top_profit=top_profit,
            sold_review_attention=sold_review_attention,
            returned_attention=returned_attention,
            aged_unsold=aged_unsold,
            negative_profit=negative_profit,
            report_viz=report_viz,
            range_key=range_key,
            trend_mode=trend_mode,
            start=start_date.isoformat() if start_date else "",
            end=end_date.isoformat() if end_date else "",
            top_n=top_n,
        )

    @app.route("/item/new", methods=["GET", "POST"])
    @auth_required
    def item_new():
        prefill = {
            "barcode": request.args.get("barcode", "").strip(),
            "item_name": request.args.get("title", "").strip(),
            "platform": request.args.get("platform", "eBay").strip(),
            "cog": request.args.get("cog", "").strip(),
            "sale_price": request.args.get("sale_price", "").strip(),
            "shipping": request.args.get("shipping", "").strip(),
            "source_location": request.args.get("source_location", "").strip(),
            "notes": request.args.get("notes", "").strip(),
        }
        prefill_barcode = prefill["barcode"]
        if request.args.get("scanner_draft") == "1":
            draft_bits = ["ScannerDraft:yes"]
            if prefill["item_name"]:
                draft_bits.append(f"ScannerSearch:{prefill['item_name']}")
            if prefill["barcode"]:
                draft_bits.append(f"ScannerBarcode:{prefill['barcode']}")
            prefill["notes"] = "\n".join([p for p in [prefill["notes"], "\n".join(draft_bits)] if p]).strip()

        if request.method == "POST":
            submitted_ebay_url = _normalize_url(request.form.get("ebay_item_url", ""))
            extracted_item_number = _extract_ebay_item_number(submitted_ebay_url)
            item = Item(
                item_name=request.form.get("item_name", "").strip(),
                category=request.form.get("category", "").strip() or None,
                sub_category=request.form.get("sub_category", "").strip() or None,
                platform=request.form.get("platform", "").strip() or None,
                notes=request.form.get("notes", "").strip() or None,
                source_location=request.form.get("source_location", "").strip() or None,
                barcode=request.form.get("barcode", "").strip() or None,
                ebay_item_number=extracted_item_number,
                ebay_item_url=_ebay_item_url(extracted_item_number) if extracted_item_number else submitted_ebay_url,
                cog=parse_float(request.form.get("cog")),
                sale_price=parse_float(request.form.get("sale_price")),
                ad_fee=parse_float(request.form.get("ad_fee")),
                ebay_fee=parse_float(request.form.get("ebay_fee")),
                shipping=parse_float(request.form.get("shipping")),
                buyer_paid_amount=parse_float(request.form.get("buyer_paid_amount")),
                date_listed=parse_date(request.form.get("date_listed")),
                date_sold=parse_date(request.form.get("date_sold")),
                sold=(request.form.get("sold") == "Y"),
            )

            if not item.item_name:
                flash("Item Name is required.", "error")
                categories = get_distinct_values(Item, Item.category)
                sub_categories = get_distinct_values(Item, Item.sub_category)
                platforms = get_distinct_values(Item, Item.platform)
                source_locations = get_distinct_values(Item, Item.source_location)

                return render_template(
                    "item_new.html",
                    item=item,
                    categories=categories,
                    sub_categories=sub_categories,
                    platforms=platforms,
                    source_locations=source_locations,
                    prefill_barcode=prefill_barcode,
                    prefill=prefill,
                )

            db.session.add(item)
            db.session.commit()  # assigns SKU

            files = request.files.getlist("photos")
            accepted_photos = 0
            for f in files:
                if not f or f.filename == "":
                    continue
                if not allowed_file(f.filename):
                    flash(f"Skipped {f.filename}: unsupported file type.", "warning")
                    continue
                if accepted_photos >= app.config["MAX_PHOTOS_PER_ITEM"]:
                    flash(f"Only the first {app.config['MAX_PHOTOS_PER_ITEM']} photos were saved.", "warning")
                    break

                safe = secure_filename(f.filename)
                ts = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
                ext = safe.rsplit(".", 1)[1].lower()
                stored_name = f"SKU{item.sku}_{ts}.{ext}"

                save_path = os.path.join(app.config["UPLOAD_FOLDER"], stored_name)
                f.save(save_path)
                if not process_image(save_path):
                    if os.path.exists(save_path):
                        os.remove(save_path)
                    flash(f"Skipped {f.filename}: file is not a valid image.", "warning")
                    continue

                db.session.add(ItemImage(item_sku=item.sku, filename=stored_name))
                accepted_photos += 1

            db.session.commit()
            flash(f"Created item SKU #{item.sku}.", "success")
            return redirect(url_for("item_detail", sku=item.sku))

        categories = get_distinct_values(Item, Item.category)
        sub_categories = get_distinct_values(Item, Item.sub_category)
        platforms = get_distinct_values(Item, Item.platform)
        source_locations = get_distinct_values(Item, Item.source_location)
        return render_template(
            "item_new.html",
            categories=categories,
            sub_categories=sub_categories,
            platforms=platforms,
            source_locations=source_locations,
            prefill_barcode=prefill_barcode,
            prefill=prefill,
        )

    @app.route("/item/<int:sku>")
    @auth_required
    def item_detail(sku: int):
        item = Item.query.get_or_404(sku)
        if _apply_note_financial_tags(item):
            db.session.commit()
        return_to = _safe_return_url(request.args.get("return_to")) or url_for("index")
        previous_item, next_item = _item_neighbors(item, return_to)
        return render_template(
            "item_detail.html",
            item=item,
            return_to=return_to,
            previous_item=previous_item,
            next_item=next_item,
        )

    @app.route("/item/<int:sku>/fetch-ebay-photo", methods=["POST"])
    @auth_required
    def fetch_ebay_photo(sku: int):
        item = Item.query.get_or_404(sku)
        raw_return_to = request.form.get("return_to")
        return_to = _safe_return_url(raw_return_to) if raw_return_to else url_for("item_detail", sku=item.sku)
        url = _normalize_url(request.form.get("ebay_item_url", "") or item.ebay_item_url or "")

        if not url:
            flash("Paste an eBay item URL first.", "error")
            return redirect(url_for("item_detail", sku=item.sku, return_to=return_to))

        is_ebay_page = _host_allowed_for_ebay_page(url)
        is_ebay_image = _host_allowed_for_ebay_image(url)

        if not is_ebay_page and not is_ebay_image:
            flash("For safety, photo fetch only accepts eBay item URLs or eBay image URLs.", "error")
            return redirect(url_for("item_detail", sku=item.sku, return_to=return_to))

        if _host_is_private_or_local(url):
            flash("That URL could not be fetched safely.", "error")
            return redirect(url_for("item_detail", sku=item.sku, return_to=return_to))

        if len(item.images or []) >= app.config["MAX_PHOTOS_PER_ITEM"]:
            flash(f"This item already has the maximum of {app.config['MAX_PHOTOS_PER_ITEM']} photos.", "warning")
            return redirect(url_for("item_detail", sku=item.sku, return_to=return_to))

        try:
            if is_ebay_image:
                image_url = url
            else:
                page_bytes, content_type = _fetch_url_bytes(url, max_bytes=2 * 1024 * 1024)
                page_text = page_bytes.decode("utf-8", errors="replace")
                image_url = _extract_og_image(page_text)
                if not image_url:
                    flash("Could not find a main photo on that eBay page.", "warning")
                    return redirect(url_for("item_detail", sku=item.sku, return_to=return_to))
                image_url = urllib.parse.urljoin(url, image_url)

            _save_image_from_url(item, image_url, app.config["UPLOAD_FOLDER"])
            if is_ebay_page:
                item.ebay_item_url = url
            db.session.commit()
            flash("Fetched the main eBay photo and added it to this item.", "success")
        except (urllib.error.URLError, TimeoutError, ValueError) as e:
            db.session.rollback()
            flash(f"Could not fetch the eBay photo: {e}", "error")
        except Exception as e:
            db.session.rollback()
            flash("Could not fetch the eBay photo. eBay may have blocked the request or changed the page.", "error")

        return redirect(url_for("item_detail", sku=item.sku, return_to=return_to))

    @app.route("/item/<int:sku>/confirm-sold", methods=["POST"])
    @auth_required
    def item_confirm_sold(sku: int):
        item = Item.query.get_or_404(sku)
        return_to = _safe_return_url(request.form.get("return_to"))
        item.sold = True
        item.sold_confirmed = True
        item.pending_shipping = True
        item.canceled = False
        db.session.commit()
        flash(f"Confirmed SKU #{item.sku} as sold. Added to Pending Shipping.", "success")
        return redirect(return_to)

    @app.route("/item/<int:sku>/mark-shipped", methods=["POST"])
    @auth_required
    def item_mark_shipped(sku: int):
        item = Item.query.get_or_404(sku)
        return_to = _safe_return_url(request.form.get("return_to"))
        item.sold = True
        item.sold_confirmed = True
        item.pending_shipping = False
        item.canceled = False
        if not item.date_shipped:
            item.date_shipped = datetime.utcnow().date()
        db.session.commit()
        flash(f"Marked SKU #{item.sku} as shipped.", "success")
        return redirect(return_to)

    @app.route("/item/<int:sku>/edit", methods=["GET", "POST"])
    @auth_required
    def item_edit(sku: int):
        item = Item.query.get_or_404(sku)
        raw_return_to = request.form.get("return_to") if request.method == "POST" else request.args.get("return_to")
        return_to = _safe_return_url(raw_return_to) if raw_return_to else url_for("item_detail", sku=item.sku)

        if request.method == "POST":
            item.item_name = request.form.get("item_name", "").strip()
            item.category = request.form.get("category", "").strip() or None
            item.sub_category = request.form.get("sub_category", "").strip() or None
            item.platform = request.form.get("platform", "").strip() or None
            item.notes = request.form.get("notes", "").strip() or None
            item.source_location = request.form.get("source_location", "").strip() or None
            item.barcode = request.form.get("barcode", "").strip() or None
            submitted_ebay_url = _normalize_url(request.form.get("ebay_item_url", ""))
            extracted_item_number = _extract_ebay_item_number(submitted_ebay_url)
            if extracted_item_number:
                item.ebay_item_number = extracted_item_number
                item.ebay_item_url = _ebay_item_url(extracted_item_number)
            elif submitted_ebay_url:
                item.ebay_item_url = submitted_ebay_url
            else:
                if not _sync_ebay_url_from_number(item):
                    item.ebay_item_url = None

            item.cog = parse_float(request.form.get("cog"))
            item.sale_price = parse_float(request.form.get("sale_price"))
            item.ad_fee = parse_float(request.form.get("ad_fee"))
            item.ebay_fee = parse_float(request.form.get("ebay_fee"))
            item.shipping = parse_float(request.form.get("shipping"))
            item.buyer_paid_amount = parse_float(request.form.get("buyer_paid_amount"))
            item.refund_amount = parse_float(request.form.get("refund_amount"))

            item.date_listed = parse_date(request.form.get("date_listed"))
            item.date_sold = parse_date(request.form.get("date_sold"))
            item.date_returned = parse_date(request.form.get("date_returned"))
            item.date_shipped = parse_date(request.form.get("date_shipped"))
            item.tracking_number = request.form.get("tracking_number", "").strip() or None
            item.return_reference_id = request.form.get("return_reference_id", "").strip() or None
            item.sold = (request.form.get("sold") == "Y")
            item.pending_shipping = (request.form.get("pending_shipping") == "Y")
            item.canceled = (request.form.get("canceled") == "Y")
            item.returned = (request.form.get("returned") == "Y")
            if item.canceled:
                item.sold = False
                item.sold_confirmed = False
                item.pending_shipping = False
            elif not item.sold:
                item.sold_confirmed = False
                item.pending_shipping = False
            else:
                item.canceled = False
                if item.tracking_number or item.date_shipped:
                    item.pending_shipping = False

            if not item.item_name:
                flash("Item Name is required.", "error")
                categories = get_distinct_values(Item, Item.category)
                sub_categories = get_distinct_values(Item, Item.sub_category)
                platforms = get_distinct_values(Item, Item.platform)
                source_locations = get_distinct_values(Item, Item.source_location)
                return render_template(
                    "item_edit.html",
                    item=item,
                    categories=categories,
                    sub_categories=sub_categories,
                    platforms=platforms,
                    source_locations=source_locations,
                    return_to=return_to,
                )

            files = request.files.getlist("photos")
            accepted_photos = 0
            for f in files:
                if not f or f.filename == "":
                    continue
                if not allowed_file(f.filename):
                    flash(f"Skipped {f.filename}: unsupported file type.", "warning")
                    continue
                if accepted_photos >= app.config["MAX_PHOTOS_PER_ITEM"]:
                    flash(f"Only the first {app.config['MAX_PHOTOS_PER_ITEM']} photos were saved.", "warning")
                    break

                safe = secure_filename(f.filename)
                ts = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
                ext = safe.rsplit(".", 1)[1].lower()
                stored_name = f"SKU{item.sku}_{ts}.{ext}"

                save_path = os.path.join(app.config["UPLOAD_FOLDER"], stored_name)
                f.save(save_path)
                if not process_image(save_path):
                    if os.path.exists(save_path):
                        os.remove(save_path)
                    flash(f"Skipped {f.filename}: file is not a valid image.", "warning")
                    continue

                db.session.add(ItemImage(item_sku=item.sku, filename=stored_name))
                accepted_photos += 1

            db.session.commit()
            flash(f"Updated SKU #{item.sku}.", "success")
            return redirect(return_to)

        categories = get_distinct_values(Item, Item.category)
        sub_categories = get_distinct_values(Item, Item.sub_category)
        platforms = get_distinct_values(Item, Item.platform)
        source_locations = get_distinct_values(Item, Item.source_location)
        return render_template(
            "item_edit.html",
            item=item,
            categories=categories,
            sub_categories=sub_categories,
            platforms=platforms,
            source_locations=source_locations,
            return_to=return_to,
        )

    @app.route("/image/<int:image_id>/delete", methods=["POST"])
    @auth_required
    def delete_image(image_id: int):
        img = ItemImage.query.get_or_404(image_id)
        sku = img.item_sku
        raw_return_to = request.form.get("return_to")
        return_to = _safe_return_url(raw_return_to) if raw_return_to else url_for("index")

        path = os.path.join(app.config["UPLOAD_FOLDER"], img.filename)
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception:
            pass

        db.session.delete(img)
        db.session.commit()
        flash("Image deleted.", "success")
        return redirect(url_for("item_detail", sku=sku, return_to=return_to))

    @app.route("/item/<int:sku>/merge", methods=["POST"])
    @auth_required
    def item_merge(sku: int):
        source = Item.query.get_or_404(sku)
        return_to = _safe_return_url(request.form.get("return_to"))
        target_sku = parse_int(request.form.get("target_sku"))

        if not target_sku:
            flash("Enter the SKU you want to keep.", "error")
            return redirect(url_for("item_detail", sku=source.sku, return_to=return_to))

        if target_sku == source.sku:
            flash("Choose a different SKU to merge into.", "error")
            return redirect(url_for("item_detail", sku=source.sku, return_to=return_to))

        target = Item.query.get(target_sku)
        if not target:
            flash(f"Could not find target SKU #{target_sku}.", "error")
            return redirect(url_for("item_detail", sku=source.sku, return_to=return_to))

        copied_fields = _copy_missing_item_fields(target, source)
        _merge_notes(target, source)
        moved_images = (
            ItemImage.query
            .filter_by(item_sku=source.sku)
            .update({"item_sku": target.sku}, synchronize_session=False)
        )

        db.session.delete(source)
        db.session.commit()

        detail = []
        if moved_images:
            detail.append(f"moved {moved_images} photo{'s' if moved_images != 1 else ''}")
        if copied_fields:
            detail.append(f"filled {len(copied_fields)} missing field{'s' if len(copied_fields) != 1 else ''}")
        suffix = f" ({', '.join(detail)})" if detail else ""
        flash(f"Merged SKU #{sku} into SKU #{target.sku}{suffix}.", "success")
        return redirect(url_for("item_detail", sku=target.sku, return_to=return_to))

    @app.route("/item/<int:sku>/delete", methods=["POST"])
    @auth_required
    def item_delete(sku: int):
        item = Item.query.get_or_404(sku)
        return_to = _safe_return_url(request.form.get("return_to"))

        for img in item.images:
            path = os.path.join(app.config["UPLOAD_FOLDER"], img.filename)
            try:
                if os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass

        db.session.delete(item)
        db.session.commit()
        flash(f"Deleted SKU #{sku}.", "success")
        return redirect(return_to)
    

    return app

if __name__ == "__main__":
    app = create_app()
    app.run(host="0.0.0.0", port=5055, debug=True)
